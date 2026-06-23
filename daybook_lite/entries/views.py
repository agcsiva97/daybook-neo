from datetime import datetime, timedelta
from decimal import Decimal
import csv
import logging
import json
import openpyxl
from functools import wraps
from openpyxl.styles import Font, Alignment, PatternFill
import markdown

from pathlib import Path
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import ensure_csrf_cookie, csrf_protect
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction as db_transaction
from django.db.models import Case, DecimalField, F, Min, Sum, Value, When
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from manager.helper import manager_helper, date_helper

from .forms import TransactionForm, TransactionEditForm, TransferForm, DenominationForm, LoanForm, LoanEditForm
from .models import Transactions, Denomination, Loan
from manager.models import Shop, Ledger, Configuration, Accounts, Type
from .helpers import transactions as transaction_helper
from manager.helper.manager_helper import log_activity
from django.contrib.auth import get_user_model

User = get_user_model()
logger = logging.getLogger(__name__)




def is_admin(user):
    return user.is_superuser or user.groups.filter(name='Admin').exists()

def is_super_admin(user):
    return user.is_superuser

def is_admin_or_staff(user):
    return user.is_superuser or user.groups.filter(name__in=['Admin', 'Staff']).exists()


# Custom decorators that raise PermissionDenied instead of redirecting to login
def admin_required(view_func):
    """Decorator that requires user to be admin or superuser. Shows 403 if not."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not is_admin(request.user):
            raise PermissionDenied('You do not have permission to access this page.')
        return view_func(request, *args, **kwargs)
    return wrapper

def super_admin_required(view_func):
    """Decorator that requires user to be superuser. Shows 403 if not."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not is_super_admin(request.user):
            raise PermissionDenied('You do not have permission to access this page.')
        return view_func(request, *args, **kwargs)
    return wrapper

def admin_or_staff_required(view_func):
    """Decorator that requires user to be admin, staff, or superuser. Shows 403 if not."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not is_admin_or_staff(request.user):
            raise PermissionDenied('You do not have permission to access this page.')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def home(request):
    today = timezone.localdate()
    gold_price = manager_helper.get_gold_price()
    silver_price = manager_helper.get_silver_price()
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    default_shop = Shop.objects.get(short_name=default_shop_short_name.value)
    
    latest_transaction = Transactions.objects.filter(
        shop = default_shop
    ).order_by('-transaction_dt').first()

    if latest_transaction:
        latest_date = timezone.localtime(latest_transaction.transaction_dt).date() if hasattr(latest_transaction.transaction_dt, 'date') else timezone.localtime(latest_transaction.transaction_dt)
    else:
        latest_date = today  

    form = TransactionForm(initial={'date': latest_date})

    # Apply permission-based filtering (same as transactions view)
    if is_admin(request.user) or is_super_admin(request.user):
        transactions = Transactions.objects.filter(transaction_dt__date=today).select_related(
            'shop', 'acc', 'acc__acc_type', 'created_by', 'updated_by'
        ).order_by('-transaction_dt')
    else:
        transactions = Transactions.objects.filter(
            transaction_dt__date=today,
            acc__is_admin_only=False
        ).select_related(
            'acc',
            'acc__acc_type',
            'shop',
            'created_by',
            'updated_by'
        ).order_by('-transaction_dt')
    daily_totals = (
        Transactions.objects.filter(transaction_dt__date=today)
        .values('shop_id')
        .annotate(
            debit_total=Coalesce(
                Sum(
                    Case(
                        When(tr_type='DEBIT', then=F('amount')),
                        default=Value(0),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            credit_total=Coalesce(
                Sum(
                    Case(
                        When(tr_type='CREDIT', then=F('amount')),
                        default=Value(0),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
    )
    totals_by_shop = {item['shop_id']: item for item in daily_totals}
    shops = Shop.objects.all().order_by('name')
    
    # Calculate opening and closing balance for each shop
    shop_balances = []
    for shop in shops:
        data = transaction_helper.get_opening_balance(shop, today)  # This will calculate and cache opening balance for the shop
        
        shop_balances.append({
            'shop': shop,
            'opening_balance': data['opening_balance'],
            'closing_balance': data['closing_balance'],
        })

    print(transactions)
    context = {
        'nav_title':'Home',
        'transactions': transactions,
        'gold_price': gold_price,
        'silver_price': silver_price,
        'shop_balances': shop_balances,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'form': form,
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
        'latest_date': latest_date,
        'default_shop': default_shop,
    }
    
    return render(request, 'entries/home.html',context)

def update_gold_price(request):
    try:
        result = manager_helper.update_gold_price()
        if result:
            messages.success(request, 'Gold price updated successfully!')
        else:
            messages.error(request, 'Failed to update gold price.')
    except (ValueError, TypeError, Decimal.InvalidOperation):
        messages.error(request, 'Invalid price value. Please enter a valid number.')
    return redirect('entries:home')

def update_silver_price(request):
    try:
        result = manager_helper.update_silver_price()
        if result:
            messages.success(request, 'Silver price updated successfully!')
        else:
            messages.error(request, 'Failed to update silver price.')
    except (ValueError, TypeError, Decimal.InvalidOperation):
        messages.error(request, 'Invalid price value. Please enter a valid number.')
    return redirect('entries:home')

@login_required
def add_entries(request):
    """Add a new transaction entry"""
    loan_form     = LoanForm()
    release_form  = LoanForm()
    today = timezone.localdate()
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    loans = Loan.objects.filter(transaction_dt__date=today, type='LOAN').order_by('-transaction_dt')
    releases = Loan.objects.filter(transaction_dt__date=today, type='RELEASE').order_by('-transaction_dt')
    shops = Shop.objects.all().order_by('short_name')
    ledgers = Ledger.objects.filter(shop__short_name = default_shop_short_name.value)
    gold_price = manager_helper.get_gold_price()
    silver_price = manager_helper.get_silver_price()

    latest_transaction = Loan.objects.filter(
        ledger__in = ledgers
    ).order_by('-transaction_dt').first()

    if latest_transaction:
        latest_date = latest_date = timezone.localtime(latest_transaction.transaction_dt).date() if hasattr(latest_transaction.transaction_dt, 'date') else timezone.localtime(latest_transaction.transaction_dt)
    else:
        latest_date = today 
    
    if request.method == 'POST':
        form = TransactionForm(request.POST)

        if form.is_valid():
            transaction = form.save(commit=False)

            if request.user.is_authenticated:
                transaction.created_by = request.user
                transaction.updated_by = request.user

            # ── Resolve transaction date + time ──────────────────────
            chosen_date = form.cleaned_data.get('date')
            chosen_time = form.cleaned_data.get('time') or timezone.localtime(timezone.now()).time()
            transaction.transaction_dt = timezone.make_aware(
                datetime.combine(chosen_date, chosen_time)
            )

            logger.info("============ Transaction Creation Started ============")
            logger.info(f"Transaction -> type=[{transaction.tr_type}] | amount=[{transaction.amount}] | date=[{transaction.transaction_dt}]")

            try:
                with db_transaction.atomic():
                    shop = Shop.objects.select_for_update().get(pk=transaction.shop_id)
                    logger.info(f"Shop -> name=[{shop.short_name}]")

                    # ── Balance check for DEBIT ───────────────────────
                    if transaction.tr_type == 'DEBIT':
                        available = transaction_helper.get_balance(shop, chosen_date)
                        logger.info(f"Balance check -> available=[{available}] | required=[{transaction.amount}]")
                        if transaction.amount > available:
                            form.add_error(None, f'Insufficient balance in {shop.short_name}. Available: {available}')
                            return render(request, 'entries/add_entries.html', {
                                'nav_title': 'Add Entries',
                                'form': form,
                                'transfer_form': TransferForm(),
                                'loan_form': LoanForm(),
                                'release_form': LoanForm(),
                                'loans': loans,
                                'releases': releases,
                                'shops': shops,
                            })

                    # ── Save transaction ──────────────────────────────
                    transaction.save()
                    manager_helper.update_account_priority(transaction.acc)
                    logger.info(f"Transaction [{transaction.id}] created -> type=[{transaction.tr_type}] | amount=[{transaction.amount}] | shop=[{shop.short_name}]")
                log_activity(request, 'CREATE', 'Transaction', transaction.id, f'Transaction created: {transaction.remarks} ({transaction.amount} {transaction.tr_type}) for {shop.short_name}', shop=shop)
                logger.info("============ Transaction Creation Completed ============")
                messages.success(request, 'Transaction added successfully!')

            except Exception as e:
                logger.error(f"Error creating transaction by [{request.user.username}]: {str(e)}", exc_info=True)
                messages.error(request, 'An error occurred while creating transaction.')
                return render(request, 'entries/add_entries.html', {
                    'nav_title': 'Add Entries',
                    'form': form,
                    'transfer_form': TransferForm(),
                    'loan_form': LoanForm(),
                    'release_form': LoanForm(),
                    'shops': shops,
                    'latest_date': latest_date,
                })

            return redirect('entries:home')

        else:
            logger.warning(f"Transaction form invalid -> errors=[{form.errors}]")

    return render(request, 'entries/add_entries.html', {
        'nav_title': 'Add Entries',
        'loans': loans,
        'releases': releases,
        'shops': shops,
        'loan_form': loan_form,
        'release_form': release_form,
        'gold_price': gold_price,
        'silver_price': silver_price,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'default_shop_short_name': default_shop_short_name,
        'all_shops': shops,
        'latest_date': latest_date,
    })


@login_required
@ensure_csrf_cookie
def transactions(request):
    # Get filter parameters
    clear_filters = request.GET.get('clear_filters').lower() if request.GET.get('clear_filters') else 'false'
    from_date = (request.GET.get('from_date') or '').strip()
    to_date = (request.GET.get('to_date') or '').strip()
    shop_filter = request.GET.get('shop')
    type_filter = request.GET.get('tr_type')
    search_query = request.GET.get('search', '')
    account_filter = [a for a in request.GET.getlist('account') if a]
    account_type_filter = request.GET.get('account_type')
    amount_value = request.GET.get('amount_value', '')
    amount_operator = request.GET.get('amount_operator', 'equals')
    # Sorting option: date_desc (default) or date_asc
    sort_option = request.GET.get('sort', 'date_desc')
    fy = request.GET.get('fin_year')

    if fy is None:
        fy = date_helper.get_current_fy_string()
    
    start_date, end_date = date_helper.get_fy_dates(fy)
    
    
    all_configs = Configuration.objects.all()
    shop = None  # Initialize shop variable for context, will be set if shop_filter is applied

    # Preserve current filter parameters for pagination and HTMX requests.
    query_params = request.GET.copy()
    query_params.pop('page', None)
    filter_query = query_params.urlencode()

    if clear_filters == 'true':
        shop_filter = 'all'

    # Note: No longer setting default from_date to allow users to see all data if desired
    
    # Base queryset (select_related used for performance)
    if is_admin(request.user) or is_super_admin(request.user):
        transactions_list = Transactions.objects.filter(
            transaction_dt__gte=start_date,
            transaction_dt__lte=end_date
        ).select_related(
                'shop', 'acc', 'acc__acc_type', 'created_by', 'updated_by'
        )
    else:
        transactions_list = Transactions.objects.filter(
            acc__is_admin_only=False,
            transaction_dt__gte=start_date,
            transaction_dt__lte=end_date
        ).select_related(
            'acc',      # Essential since you are filtering/displaying account info 
            'acc__acc_type',
            'shop', 
            'created_by', 
            'updated_by'
        )

    if not shop_filter:
        default_shop_short_name = Configuration.get_value(Configuration.Key.DEFAULT_SHOP, default='')
        if default_shop_short_name:
            default_shop = Shop.objects.filter(short_name=default_shop_short_name).first()
            if default_shop:
                shop_filter = str(default_shop.id)

    # Apply ordering based on sort option (default: date descending)
    if sort_option == 'date_asc':
        transactions_list = transactions_list.order_by('transaction_dt')
    else:
        transactions_list = transactions_list.order_by('-transaction_dt')
    

    # Apply filters
    if from_date:
        from_date_obj = date_helper.parse_date_string(from_date)
        if from_date_obj:
            from_date_dt = timezone.make_aware(
                datetime.combine(from_date_obj, datetime.min.time()),
                timezone=timezone.get_current_timezone()
            )
            transactions_list = transactions_list.filter(transaction_dt__gte=from_date_dt)
    
    if to_date:
        to_date_obj = date_helper.parse_date_string(to_date)
        if to_date_obj:
            to_date_dt = timezone.make_aware(
                datetime.combine(to_date_obj, datetime.max.time()),
                timezone=timezone.get_current_timezone()
            )
            transactions_list = transactions_list.filter(transaction_dt__lte=to_date_dt)
    
    if shop_filter != 'all':
        transactions_list = transactions_list.filter(shop_id=shop_filter)
        shop = Shop.objects.filter(pk=shop_filter).first()
    
    if type_filter and type_filter in ['DEBIT', 'CREDIT']:
        transactions_list = transactions_list.filter(tr_type=type_filter)
    
    if account_filter:
        transactions_list = transactions_list.filter(acc_id__in=account_filter)
    
    if account_type_filter:
        transactions_list = transactions_list.filter(acc__acc_type_id=account_type_filter)
    
    # Apply amount filter with operator
    if amount_value:
        try:
            amount_val = Decimal(amount_value)
            if amount_operator == 'greater':
                transactions_list = transactions_list.filter(amount__gt=amount_val)
            elif amount_operator == 'lesser':
                transactions_list = transactions_list.filter(amount__lt=amount_val)
            elif amount_operator == 'equals':
                transactions_list = transactions_list.filter(amount=amount_val)
        except (ValueError, TypeError):
            pass
    
    if search_query:
        transactions_list = transactions_list.filter(remarks__icontains=search_query)
    
    # Pagination
    paginator = Paginator(transactions_list, 25)  # Show 25 transactions per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate totals for all filtered transactions (not just current page)
    totals = transactions_list.aggregate(
        debit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='DEBIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        credit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='CREDIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )
    
    # Get all shops for filter dropdown
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    
    # Get all accounts for filter dropdown
    all_accounts = Accounts.objects.all().select_related('acc_type').order_by('e_name')
    
    # Get all account types for filter dropdown
    all_account_types = Type.objects.all().order_by('e_name')
    
    if request.headers.get('HX-Request'):
        return render(request, 'entries/partials/transaction_rows.html', {
            'page_obj': page_obj,
            'is_super_admin': request.user.is_superuser,
            'is_admin': is_admin(request.user),
            'is_admin_user': is_admin(request.user) or request.user.is_superuser,
            'filter_query': filter_query,
            'sort': sort_option,
            'shop_filter': shop_filter,
        })
    
    configs={}
    for config in all_configs:
        if 'TRANS_' in config.key:
            configs[config.key] = config.value
            print(f"Config: {config.key} = {config.value}")

    context = {
        'nav_title': 'Transactions',
        'page_obj': page_obj,
        'all_transactions': transactions_list,  # All filtered transactions for printing
        'all_shops': all_shops,
        'all_accounts': all_accounts,
        'all_account_types': all_account_types,
        'from_date': from_date,
        'to_date': to_date,
        'shop_filter': shop_filter,
        'type_filter': type_filter,
        'account_filter': account_filter,
        'clear_filters': clear_filters,
        'account_type_filter': account_type_filter,
        'amount_value': amount_value,
        'amount_operator': amount_operator,
        'search_query': search_query,
        'sort': sort_option,
        'debit_total': totals['debit_total'],
        'credit_total': totals['credit_total'],
        'net_balance': totals['credit_total'] - totals['debit_total'],
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'configs': configs,
        'shop': shop,
        'filter_query': filter_query,
        'is_admin_user': is_admin(request.user) or request.user.is_superuser,
        'default_shop_short_name': default_shop_short_name,
        'fy':fy,
    }
    return render(request, 'entries/transactions.html', context)


@login_required
def transactions_print(request):
    # Use same filters as transactions but return all matching data for reporting
    shop_filter = request.GET.get('shop')
    account_filter = [a for a in request.GET.getlist('account') if a]
    fy = request.GET.get('fin_year')
    one_account = False
    account = None
    if len(account_filter) == 1:
        one_account = True
        account = Accounts.objects.get(id=account_filter[0])
    if fy is None:
        fy = date_helper.get_current_fy_string()
    start_date, end_date = date_helper.get_fy_dates(fy)
    all_configs = Configuration.objects.all()
    configs={}
    for config in all_configs:
        if 'TRANS_' in config.key:
            configs[config.key] = config.value
            print(f"Config: {config.key} = {config.value}")
    transactions = _get_filtered_transactions(request)

    shop = None
    if shop_filter:
        shop = Shop.objects.filter(pk=shop_filter).first()

    if transactions.count() > 2000:
        messages.error(request, 'Filter transaction less than 1000')
        return redirect('entries:transactions')
    
    context = {
        'transactions': transactions,
        'from_date': request.GET.get('from_date', ''),
        'to_date': request.GET.get('to_date', ''),
        'shop_filter': shop_filter,
        'shop': shop,
        'type_filter': request.GET.get('type', ''),
        'account_filter': request.GET.getlist('account', ''),
        'account_type_filter': request.GET.get('account_type', ''),
        'amount_value': request.GET.get('amount_value', ''),
        'amount_operator': request.GET.get('amount_operator', 'equals'),
        'search_query': request.GET.get('search', ''),
        'configs': configs,
        'one_account': one_account,
        'account': account,
    }
    return render(request, 'entries/transactions_print.html', context)


def _get_filtered_transactions(request):
    """Helper function to get filtered transactions for exports"""
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    shop_filter = request.GET.get('shop')
    type_filter = request.GET.get('type')
    search_query = request.GET.get('search', '')
    account_filter = [a for a in request.GET.getlist('account') if a]
    account_type_filter = request.GET.get('account_type')
    amount_value = request.GET.get('amount_value', '')
    amount_operator = request.GET.get('amount_operator', 'equals')
    sort_option = request.GET.get('sort', 'date_desc')
    fy = request.GET.get('fin_year')
    if fy is None:
        fy = date_helper.get_current_fy_string()
    start_date, end_date = date_helper.get_fy_dates(fy)
    
    # Base queryset - order by created date descending
    if is_admin(request.user) or is_super_admin(request.user):
        transactions_list = Transactions.objects.filter(
            transaction_dt__gte=start_date,
            transaction_dt__lte=end_date
        ).select_related(
                'shop', 'acc', 'acc__acc_type', 'created_by', 'updated_by'
        )
    else:
        transactions_list = Transactions.objects.filter(
            acc__is_admin_only=False,
            transaction_dt__gte=start_date,
            transaction_dt__lte=end_date
        ).select_related(
            'acc',      
            'acc__acc_type',
            'shop', 
            'created_by', 
            'updated_by'
        )
    
    # Apply filters
    if from_date:
        from_date_obj = date_helper.parse_date_string(from_date)
        if from_date_obj:
            from_date_dt = timezone.make_aware(
                datetime.combine(from_date_obj, datetime.min.time()),
                timezone=timezone.get_current_timezone()
            )
            transactions_list = transactions_list.filter(transaction_dt__gte=from_date_dt)
    
    if to_date:
        to_date_obj = date_helper.parse_date_string(to_date)
        if to_date_obj:
            to_date_dt = timezone.make_aware(
                datetime.combine(to_date_obj, datetime.max.time()),
                timezone=timezone.get_current_timezone()
            )
            transactions_list = transactions_list.filter(transaction_dt__lte=to_date_dt)
    
    if shop_filter:
        transactions_list = transactions_list.filter(shop_id=shop_filter)
    
    if type_filter and type_filter in ['DEBIT', 'CREDIT']:
        transactions_list = transactions_list.filter(tr_type=type_filter)
    
    if account_type_filter:
        transactions_list = transactions_list.filter(acc__acc_type_id=account_type_filter)
    
    # Apply amount filter with operator
    if amount_value:
        try:
            amount_val = Decimal(amount_value)
            if amount_operator == 'greater':
                transactions_list = transactions_list.filter(amount__gt=amount_val)
            elif amount_operator == 'lesser':
                transactions_list = transactions_list.filter(amount__lt=amount_val)
            elif amount_operator == 'equals':
                transactions_list = transactions_list.filter(amount=amount_val)
        except (ValueError, TypeError):
            pass
    
    if search_query:
        transactions_list = transactions_list.filter(remarks__icontains=search_query)
    
    if account_filter:
        if sort_option == 'date_asc':
            transactions_list = transactions_list.filter(acc_id__in=account_filter).order_by('acc__t_name','transaction_dt')
        else:
            transactions_list = transactions_list.filter(acc_id__in=account_filter).order_by('acc__t_name','-transaction_dt')
    else:
        if sort_option == 'date_asc':
            transactions_list = transactions_list.order_by('transaction_dt')
        else:
            transactions_list = transactions_list.order_by('-transaction_dt')

    return transactions_list

@login_required
def export_transactions_csv(request):
    transactions = _get_filtered_transactions(request)
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    shop_filter = request.GET.get('shop')
    type_filter = request.GET.get('type')
    search_query = request.GET.get('search')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.localdate()}.csv"'
    
    writer = csv.writer(response)
    
    # Add filter information
    writer.writerow(['Transactions Export'])
    writer.writerow(['Export Date:', timezone.localdate()])
    writer.writerow(['Exported By:', f'{request.user.first_name} {request.user.last_name}'])
    writer.writerow([])
    
    # Add applied filters
    writer.writerow(['Applied Filters:'])
    if from_date:
        writer.writerow(['From Date:', from_date])
    if to_date:
        writer.writerow(['To Date:', to_date])
    if shop_filter:
        try:
            shop = Shop.objects.get(pk=shop_filter)
            writer.writerow(['Shop:', shop.name])
        except Shop.DoesNotExist:
            pass
    if type_filter:
        writer.writerow(['Type:', type_filter])
    if search_query:
        writer.writerow(['Remarks Search:', search_query])
    if not any([from_date, to_date, shop_filter, type_filter, search_query]):
        writer.writerow(['No filters applied - showing all transactions'])
    writer.writerow([])
    
    # Transaction headers
    writer.writerow(['Transaction Date', 'Shop', 'Account', 'Remarks', 'Debit', 'Credit'])
    
    for transaction in transactions:
        # Combine date and time into Transaction Date
        transaction_datetime = transaction.transaction_dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Get account name (t_name), default to '-' if not available
        account_name = transaction.acc.t_name if transaction.acc else '-'
        
        # Separate debit and credit based on tr_type
        if transaction.tr_type == 'DEBIT':
            debit_amount = transaction.amount
            credit_amount = ''
        else:  # CREDIT
            debit_amount = ''
            credit_amount = transaction.amount
        
        writer.writerow([
            transaction_datetime,
            transaction.shop.name,
            account_name,
            transaction.remarks or '-',
            debit_amount,
            credit_amount,
        ])
    
    # Add totals
    totals = transactions.aggregate(
        debit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='DEBIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        credit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='CREDIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )
    
    writer.writerow([])
    writer.writerow(['', '', '', 'TOTAL', totals['debit_total'], totals['credit_total']])
    
    logger.info(f"Transactions CSV export by {request.user.username}")
    
    return response

@login_required
def export_transactions_excel(request):
    transactions = _get_filtered_transactions(request)
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    shop_filter = request.GET.get('shop')
    type_filter = request.GET.get('type')
    search_query = request.GET.get('search')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Header styling
    header_fill = PatternFill(start_color="4A7766", end_color="4A7766", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Update merged cell for new column count (6 columns now)
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Transactions Export - {timezone.localdate()}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add exported by info
    ws.merge_cells('A2:F2')
    export_cell = ws['A2']
    export_cell.value = f"Exported by: {request.user.first_name} {request.user.last_name}"
    export_cell.alignment = Alignment(horizontal="right")
    export_cell.font = Font(italic=True, size=9)
    
    current_row = 4
    
    # Add filter information
    if any([from_date, to_date, shop_filter, type_filter, search_query]):
        ws.merge_cells(f'A{current_row}:F{current_row}')
        filter_title = ws.cell(row=current_row, column=1)
        filter_title.value = "Applied Filters:"
        filter_title.font = Font(bold=True, size=11)
        current_row += 1
        
        if from_date:
            ws.cell(row=current_row, column=1, value="From Date:")
            ws.cell(row=current_row, column=2, value=from_date)
            current_row += 1
        
        if to_date:
            ws.cell(row=current_row, column=1, value="To Date:")
            ws.cell(row=current_row, column=2, value=to_date)
            current_row += 1
        
        if shop_filter:
            try:
                shop_obj = Shop.objects.get(pk=shop_filter)
                # ws.cell(row=current_row, column=1, value="Shop:")
                # ws.cell(row=current_row, column=2, value=shop_obj.name)
                current_row += 1
            except Shop.DoesNotExist:
                pass
        
        if type_filter:
            ws.cell(row=current_row, column=1, value="Type:")
            type_cell = ws.cell(row=current_row, column=2, value=type_filter)
            if type_filter == 'DEBIT':
                type_cell.font = Font(color="FF0000", bold=True)
            else:
                type_cell.font = Font(color="008000", bold=True)
            current_row += 1
        
        if search_query:
            ws.cell(row=current_row, column=1, value="Remarks Search:")
            ws.cell(row=current_row, column=2, value=search_query)
            current_row += 1
        
        current_row += 1  # Add space after filters
    else:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        no_filter_cell = ws.cell(row=current_row, column=1)
        no_filter_cell.value = "No filters applied - showing all transactions"
        no_filter_cell.font = Font(italic=True, size=9)
        current_row += 2
    
    # Headers
    if not shop_filter:
        headers = ['Transaction Date', 'Shop', 'Account', 'Remarks', 'Debit', 'Credit']
    else:
        headers = ['Transaction Date', 'Account', 'Remarks', 'Debit', 'Credit']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    current_row += 1
    
    # Transaction Data
    for transaction in transactions:
        # Combine date and time into Transaction Date
        transaction_datetime = transaction.transaction_dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Get account name (t_name), default to '-' if not available
        account_name = transaction.acc.t_name if transaction.acc else '-'
        
        # Separate debit and credit based on tr_type
        if transaction.tr_type == 'DEBIT':
            debit_amount = float(transaction.amount)
            credit_amount = None
        else:  # CREDIT
            debit_amount = None
            credit_amount = float(transaction.amount)
        
        ws.cell(row=current_row, column=1, value=transaction_datetime)
        if not shop_filter:
            ws.cell(row=current_row, column=2, value=transaction.shop.name)
            ws.cell(row=current_row, column=3, value=account_name)
            ws.cell(row=current_row, column=4, value=transaction.remarks or '-')
        
            # Debit column
            debit_cell = ws.cell(row=current_row, column=5, value=debit_amount)
            if debit_amount is not None:
                debit_cell.font = Font(color="FF0000")
            
            # Credit column
            credit_cell = ws.cell(row=current_row, column=6, value=credit_amount)
            if credit_amount is not None:
                credit_cell.font = Font(color="008000")
        else:
            ws.cell(row=current_row, column=2, value=account_name)
            ws.cell(row=current_row, column=3, value=transaction.remarks or '-')
        
            # Debit column
            debit_cell = ws.cell(row=current_row, column=4, value=debit_amount)
            if debit_amount is not None:
                debit_cell.font = Font(color="FF0000")
            
            # Credit column
            credit_cell = ws.cell(row=current_row, column=5, value=credit_amount)
            if credit_amount is not None:
                credit_cell.font = Font(color="008000")
        
        current_row += 1
    
    # Totals
    totals = transactions.aggregate(
        debit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='DEBIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        credit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='CREDIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )
    
    total_row = current_row + 1
    total_cell = ws.cell(row=total_row, column=4, value='TOTAL')
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal="right")
    
    debit_total_cell = ws.cell(row=total_row, column=5, value=float(totals['debit_total']))
    debit_total_cell.font = Font(bold=True, color="FF0000")
    
    credit_total_cell = ws.cell(row=total_row, column=6, value=float(totals['credit_total']))
    credit_total_cell.font = Font(bold=True, color="008000")

    net_balance = totals['credit_total'] - totals['debit_total']
    balance_cell = ws.cell(row=total_row + 1, column=4, value='NET BALANCE')
    balance_cell.font = Font(bold=True)
    balance_cell.alignment = Alignment(horizontal="right")
    if net_balance < 0:
        net_balance_cell = ws.cell(row=total_row + 1, column=5, value=float(net_balance))
        net_balance_cell.font = Font(bold=True, color="FF0000")
    else:
        net_balance_cell = ws.cell(row=total_row + 1, column=6, value=float(net_balance))
        net_balance_cell.font = Font(bold=True, color="008000")

    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Transaction Date
    ws.column_dimensions['B'].width = 20  # Shop
    ws.column_dimensions['C'].width = 20  # Account
    ws.column_dimensions['D'].width = 25  # Remarks
    ws.column_dimensions['E'].width = 12  # Debit
    ws.column_dimensions['F'].width = 12  # Credit
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="transactions_{timezone.localdate()}.xlsx"'
    wb.save(response)
    
    logger.info(f"Transactions Excel export by {request.user.username}")
    
    return response

@login_required
@admin_or_staff_required
def edit_transaction(request, pk):
    """Edit an existing transaction"""
    transaction = get_object_or_404(Transactions, pk=pk)
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    all_shops = Shop.objects.all()

    # ── Snapshot old values ──────────────────────────────────────────
    old_amount  = transaction.amount
    old_type    = transaction.tr_type
    old_shop    = transaction.shop
    old_acc     = transaction.acc
    old_date    = timezone.localtime(transaction.transaction_dt).date()

    if request.method == 'POST':
        form = TransactionEditForm(request.POST, instance=transaction, user=request.user)

        if form.is_valid():
            updated_transaction = form.save(commit=False)

            if request.user.is_authenticated:
                updated_transaction.updated_by = request.user

            # ── Resolve new transaction date + time ──────────────────
            chosen_date = form.cleaned_data.get('date')
            chosen_time = form.cleaned_data.get('time') or timezone.localtime(timezone.now()).time()
            updated_transaction.transaction_dt = timezone.make_aware(
                datetime.combine(chosen_date, chosen_time)
            )
            new_date = chosen_date
            new_shop = updated_transaction.shop
            new_acc  = updated_transaction.acc
            new_type = updated_transaction.tr_type
            new_amount = updated_transaction.amount

            logger.info("============ Transaction Update Started ============")
            logger.info(f"Old -> amount=[{old_amount}] | type=[{old_type}] | shop=[{old_shop}] | account=[{old_acc}] | date=[{old_date}]")
            logger.info(f"New -> amount=[{new_amount}] | type=[{new_type}] | shop=[{new_shop}] | account=[{new_acc}] | date=[{new_date}]")

            try:
                with db_transaction.atomic():

                    # ── Balance check for DEBIT ───────────────────────
                    if new_type == 'DEBIT':
                        available = transaction_helper.get_balance(new_shop, new_date)
                        # Add back old amount if same shop to avoid double counting
                        if old_shop.id == new_shop.id and old_type == 'DEBIT':
                            available += old_amount
                        logger.info(f"Balance check -> shop=[{new_shop.short_name}] | available=[{available}] | required=[{new_amount}]")
                        if new_amount > available:
                            form.add_error(None, f'Insufficient balance in {new_shop.short_name} on {new_date}. Available: {available}')
                            return render(request, 'entries/edit-transaction.html', {
                                'nav_title': 'Other Transactions',
                                'form': form,
                                'transaction': transaction,
                                'is_super_admin': request.user.is_superuser,
                                'is_admin': is_admin(request.user),
                            })

                    # ── Save updated transaction ──────────────────────
                    updated_transaction.save()
                    logger.info(f"Transaction [{pk}] updated successfully")

                # ── Build change summary ──────────────────────────────
                changes = []
                if old_amount != new_amount:
                    changes.append(f"amount: [{old_amount}] -> [{new_amount}]")
                if old_type != new_type:
                    changes.append(f"type: [{old_type}] -> [{new_type}]")
                if old_shop.id != new_shop.id:
                    changes.append(f"shop: [{old_shop.short_name}] -> [{new_shop.short_name}]")
                old_acc_id = old_acc.id if old_acc else None
                new_acc_id = new_acc.id if new_acc else None
                if old_acc_id != new_acc_id:
                    changes.append(f"account: [{old_acc.t_name if old_acc else 'None'}] -> [{new_acc.t_name if new_acc else 'None'}]")
                if old_date != new_date:
                    changes.append(f"date: [{old_date}] -> [{new_date}]")
                logger.info(f"Transaction [{pk}] changes -> {', '.join(changes) if changes else 'no changes'}")
                logger.info("============ Transaction Update Completed ============")
                messages.success(request, 'Transaction updated successfully!')

            except Exception as e:
                logger.error(f"Error editing transaction [{pk}] by [{request.user.username}]: {str(e)}", exc_info=True)
                messages.error(request, 'An error occurred while editing transaction.')
                return render(request, 'entries/edit-transaction.html', {
                    'nav_title': 'Other Transactions',
                    'form': form,
                    'transaction': transaction,
                    'is_super_admin': request.user.is_superuser,
                    'is_admin': is_admin(request.user),
                })
            log_activity(request, 'UPDATE', 'Transaction', transaction.id, f'Transaction updated: {transaction.remarks} ({transaction.amount} {transaction.tr_type}) for {transaction.shop.short_name}', shop=transaction.shop)
            return redirect('entries:transactions')

        else:
            logger.warning(f"Transaction form invalid -> errors=[{form.errors}]")

    else:
        form = TransactionEditForm(
            instance=transaction,
            user=request.user,
            initial={
                'date': timezone.localtime(transaction.transaction_dt).date(),
                'time': timezone.localtime(transaction.transaction_dt).time().replace(second=0, microsecond=0),
            }
        )

    return render(request, 'entries/edit-transaction.html', {
        'nav_title': 'Other Transactions',
        'form': form,
        'transaction': transaction,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    })

@login_required
@csrf_protect
@admin_required
def delete_transaction(request, pk):
    """Delete a transaction (admin only)"""
    transaction = get_object_or_404(Transactions, pk=pk)

    if request.method == 'POST':
        # ✅ Read next URL from POST body, fallback to default
        next_url = request.POST.get('next') or 'entries:transactions'

        transaction_id   = transaction.id
        transaction_amount = transaction.amount
        transaction_type = transaction.tr_type
        shop_name        = transaction.shop.short_name
        transaction_date = timezone.localtime(transaction.transaction_dt).date()

        logger.info("============ Transaction Deletion Started ============")
        logger.info(f"Transaction -> id=[{transaction_id}] | type=[{transaction_type}] | amount=[{transaction_amount}] | shop=[{shop_name}] | date=[{transaction_date}]")

        try:
            # Option 1 — warn user if transaction is loan-linked
            with db_transaction.atomic():
                if transaction_helper.is_loan_transaction(transaction.loan_tr_type):
                    messages.error(request, 'This transaction is linked to a loan entry. Please delete it from Loan Transactions page instead.')
                    return redirect('entries:transactions')
                log_activity(request, 'DELETE', 'Transaction', transaction.id, f'Transaction deleted: {transaction.remarks} ({transaction.amount} {transaction.tr_type}) for {transaction.shop.short_name}', shop=transaction.shop)
                transaction.delete()
                logger.warning(f"Transaction deleted by [{request.user.username}] -> id=[{transaction_id}] | type=[{transaction_type}] | amount=[{transaction_amount}] | shop=[{shop_name}]")
            messages.success(request, 'Transaction deleted successfully!')
            logger.info("============ Transaction Deletion Completed ============")

        except Exception as e:
            logger.error(f"Error deleting transaction [{transaction_id}] by [{request.user.username}]: {str(e)}", exc_info=True)
            messages.error(request, 'An error occurred while deleting transaction.')

        return redirect(next_url)

    return render(request, 'entries/delete-transaction.html', {
        'nav_title': 'Other Transactions',
        'transaction': transaction,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
    })


@login_required
def report(request):
    # Get filter parameters
    report_date = request.GET.get('date')
    shop_id = request.GET.get('shop')
    all_configs = Configuration.objects.all()
    
    # Use today's date if no date specified
    if not report_date:
        report_date = timezone.localdate()
    else:
        # Convert string date to date object for template rendering
        try:
            report_date = datetime.strptime(report_date, '%Y-%m-%d').date()
        except ValueError:
            report_date = timezone.localdate()
    
    next_day = report_date + timedelta(days=1)
    prev_day = report_date - timedelta(days=1)
    
    # Log report generation
    filter_info = f"shop_id={shop_id}" if shop_id else "all shops"
    logger.info(f"Report generated by {request.user.username}: date={report_date}, filter={filter_info}")
    
    # Base queryset - filter by date
    if is_admin(request.user) or is_super_admin(request.user):
        transactions = Transactions.objects.filter(
            transaction_dt__date=report_date
        ).select_related(
                'shop', 'acc', 'acc__acc_type', 'created_by', 'updated_by'
        )
    else:
        transactions = Transactions.objects.filter(
            transaction_dt__date=report_date,
            acc__is_admin_only=False
        ).select_related(
            'acc',      # Essential since you are filtering/displaying account info 
            'acc__acc_type',
            'shop', 
            'created_by', 
            'updated_by'
        )
        
    # Apply shop filter if specified
    if not shop_id:
        default_shop_short_name = Configuration.get_value(Configuration.Key.DEFAULT_SHOP, default='')
        if default_shop_short_name:
            default_shop = Shop.objects.filter(short_name=default_shop_short_name).first()
            if default_shop:
                shop_id = str(default_shop.id)

    transactions = transactions.filter(shop_id=shop_id)
    
    transactions = transactions.order_by('transaction_dt')
    
    # Calculate totals for the day
    totals = transactions.aggregate(
        debit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='DEBIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        ),
        credit_total=Coalesce(
            Sum(
                Case(
                    When(tr_type='CREDIT', then=F('amount')),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )
    
    # Calculate shop balances (opening and closing)
    shop_summaries = []
    shops_to_process = Shop.objects.filter(id=shop_id) if shop_id else Shop.objects.all()
    
    for shop in shops_to_process.order_by('name'):
        data = transaction_helper.get_opening_balance(shop, report_date)
        
        # Get denominations for this shop on this date
        all_denoms = Denomination.objects.filter(
            shop=shop,
            denomination_dt=report_date
        )
        
        # Find the single latest denomination group by (denomination_group_order DESC, created_at DESC)
        latest_denom = all_denoms.order_by('-denomination_group_order', '-created_at').first()
        
        if latest_denom:
            # Sum only amounts belonging to that latest key
            denom_total = all_denoms.filter(key=latest_denom.key).aggregate(
                total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField(max_digits=12, decimal_places=2))
            )['total']
        else:
            denom_total = Decimal('0.00')
        
        if denom_total > 0:
            balance = data['closing_balance'] - denom_total
        else:
            balance = 0
        
        shop_summaries.append({
            'shop': shop,
            'opening_balance': data['opening_balance'],
            'closing_balance': data['closing_balance'],
            'debit_total': data['day_total_debit'],
            'credit_total': data['day_total_credit'],
            'denomination_total': denom_total,
            'balance': balance,
        })
    
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    
    # Get loan and release summaries for the report date
    loan_release_filter = {'transaction_dt__date': report_date}
    if shop_id:
        loan_release_filter['shop_id'] = shop_id

    loan_entries = Loan.objects.filter(
        type='LOAN', **loan_release_filter
    ).select_related('shop').order_by('transaction_dt')

    release_entries = Loan.objects.filter(
        type='RELEASE', **loan_release_filter
    ).select_related('shop').order_by('transaction_dt')

    # Calculate totals for loans
    loan_totals = Loan.objects.filter(
        type='LOAN', **loan_release_filter
    ).aggregate(
        total_principal=Coalesce(Sum('principal'), Value(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
        total_interest=Coalesce(Sum('interest'), Value(0), output_field=DecimalField(max_digits=12, decimal_places=2))
    )

    # Calculate totals for releases
    release_totals = Loan.objects.filter(
        type='RELEASE', **loan_release_filter
    ).aggregate(
        total_principal=Coalesce(Sum('principal'), Value(0), output_field=DecimalField(max_digits=12, decimal_places=2)),
        total_interest=Coalesce(Sum('interest'), Value(0), output_field=DecimalField(max_digits=12, decimal_places=2))
    )
    
    # Fetch denominations for the report date, grouped by key and ordered by denomination_group_order and denomination_order
    if is_admin(request.user) or is_super_admin(request.user):
        denom_filter = {'denomination_dt': report_date}
    else:
        denom_filter = {'created_by': request.user, 'denomination_dt': report_date}
    
    if shop_id:
        denom_filter['shop_id'] = shop_id
    
    # Query all denominations for the date and group them by key
    denomination_entries_qs = (
        Denomination.objects
        .filter(**denom_filter)
        .select_related('shop', 'created_by')
        .order_by('key', 'denomination_group_order', 'denomination_order')
    )
    
    # Group denominations by key, maintaining order of denomination_group_order and denomination_order
    denomination_by_key = {}
    for entry in denomination_entries_qs:
        key = entry.key
        
        if key not in denomination_by_key:
            denomination_by_key[key] = {
                'key': key,
                'time_period': entry.time_period,
                'shop_name': entry.shop.name if entry.shop else '',
                'user': f"{entry.created_by.first_name or ''} {entry.created_by.last_name or ''}".strip() if entry.created_by else '',
                'groups': {},  # Dictionary to hold groups by denomination_group_order
                'total': Decimal('0.00'),
            }
        
        # Group within key by denomination_group_order
        group_order = entry.denomination_group_order
        if group_order not in denomination_by_key[key]['groups']:
            denomination_by_key[key]['groups'][group_order] = []
        
        # Add denomination entry to the group
        denomination_by_key[key]['groups'][group_order].append({
            'denomination': entry.denomination,
            'count': entry.count,
            'amount': entry.amount,
            'denomination_order': entry.denomination_order,
        })
        denomination_by_key[key]['total'] += entry.amount
    
    # Convert to a sorted list structure for template rendering
    # Sort by key, then groups by denomination_group_order
    denomination_periods = []
    for key, key_data in sorted(denomination_by_key.items()):
        # Convert groups dict to sorted list of (group_order, rows) tuples
        sorted_groups = sorted(key_data['groups'].items(), key=lambda x: x[0])
        key_data['groups'] = sorted_groups
        denomination_periods.append((key, key_data))
    
    configs={}
    for config in all_configs:
        if 'D_REP' in config.key:
            configs[config.key] = config.value
            print(f"Config: {config.key} = {config.value}")

    context = {
        'nav_title': 'Report',
        'transactions': transactions,
        'all_shops': all_shops,
        'configs': configs,
        'report_date': report_date,
        'selected_shop': shop_id,
        'debit_total': round(totals['debit_total'], 2),
        'credit_total': round(totals['credit_total'], 2),
        'shop_summaries': shop_summaries,
        'loan_entries': loan_entries,
        'release_entries': release_entries,
        'loan_totals': loan_totals,
        'release_totals': release_totals,
        'denomination_periods': denomination_periods,
        'next_day': next_day,
        'prev_day': prev_day,
        'is_super_admin': request.user.is_superuser,
        'default_shop_short_name':default_shop_short_name,
    }
    return render(request, 'entries/report.html', context)


@login_required
def denomination(request):
    if request.method == 'POST':
        form = DenominationForm(request.POST)
        if form.is_valid():
            # Get form data
            time_period = form.cleaned_data.get('time_period')
            chosen_date = form.cleaned_data.get('date') or timezone.localdate()
            chosen_time = form.cleaned_data.get('time') or timezone.localtime(timezone.now()).time()
            denomination_dt = timezone.make_aware(
                datetime.combine(chosen_date, chosen_time)
            )
            
            # Generate key: DDMMYYYY-XX-Username
            shop = form.cleaned_data.get('shop')
            print(shop)
            time_period_code = {
                'MORNING': '01',
                'AFTERNOON': '02',
                'EVENING': '03',
                'NIGHT': '04'
            }.get(time_period, '00')
            key = f"{shop.short_name}-{denomination_dt.strftime('%d%m%Y')}-{time_period_code}-{request.user.username}"
            
            # Check if key already exists
            if Denomination.objects.filter(key=key).exists():
                messages.error(request, f'Denomination for {time_period.title()} on {denomination_dt.strftime("%d-%m-%Y")} already exists!')
                return render(request, 'entries/denomination.html', {'form': form})
            
            note_2000 = form.cleaned_data.get('note_2000') or 0
            note_500 = form.cleaned_data.get('note_500') or 0
            note_200 = form.cleaned_data.get('note_200') or 0
            note_100 = form.cleaned_data.get('note_100') or 0
            note_50 = form.cleaned_data.get('note_50') or 0
            note_20 = form.cleaned_data.get('note_20') or 0
            note_10 = form.cleaned_data.get('note_10') or 0
            coins = form.cleaned_data.get('coins') or Decimal('0.00')
            damage = form.cleaned_data.get('damage') or Decimal('0.00')
            inside = form.cleaned_data.get('inside') or Decimal('0.00')
            bundle_500 = form.cleaned_data.get('bundle_500') or 0
            bundle_200 = form.cleaned_data.get('bundle_200') or 0
            bundle_100 = form.cleaned_data.get('bundle_100') or 0
            bundle_50 = form.cleaned_data.get('bundle_50') or 0
            bundle_20 = form.cleaned_data.get('bundle_20') or 0
            bundle_10 = form.cleaned_data.get('bundle_10') or 0
            
            try:
                # Calculate amounts and create denomination records
                denominations = [
                    ('2000', note_2000, note_2000 * 2000,1),
                    ('500', note_500, note_500 * 500,2),
                    ('200', note_200, note_200 * 200,3),
                    ('100', note_100, note_100 * 100,4),
                    ('50', note_50, note_50 * 50,5),
                    ('20', note_20, note_20 * 20,6),
                    ('10', note_10, note_10 * 10,7),
                    ('Coins', 1 if coins > 0 else 0, coins,8),
                    ('Damage', 1 if damage > 0 else 0, damage,9),
                    ('Inside', 1 if inside > 0 else 0, inside,10),
                    ('500 Bundle', bundle_500, bundle_500 * 500 * 100,11),
                    ('200 Bundle', bundle_200, bundle_200 * 200 * 100,12),
                    ('100 Bundle', bundle_100, bundle_100 * 100 * 100,13),
                    ('50 Bundle', bundle_50, bundle_50 * 50 * 100,14),
                    ('20 Bundle', bundle_20, bundle_20 * 20 * 100,15),
                    ('10 Bundle', bundle_10, bundle_10 * 10 * 100,16),
                ]
                
                for denom_name, count, amount, order in denominations:
                    if count > 0 or amount > 0:
                        Denomination.objects.create(
                            denomination_dt = denomination_dt,
                            denomination=denom_name,
                            count=count,
                            amount=Decimal(str(amount)),
                            time_period=time_period,
                            key=key,
                            shop=shop,
                            created_by=request.user,
                            updated_by=request.user,
                            denomination_order=order,
                            denomination_group_order=time_period_code,
                        )
                
                log_activity(request, 'CREATE', 'Denomination', key, f'Denomination created: {key}', shop=shop)
                logger.info(f"Denomination added by {request.user.username}")
                messages.success(request, 'Denomination added successfully!')
                transaction_helper.purge_old_denominations()
                return redirect('entries:denominations')
            except Exception as e:
                logger.error(f"Error adding denomination by {request.user.username}: {str(e)}", exc_info=True)
                messages.error(request, 'An error occurred while adding denomination.')
        else:
            messages.error(request, 'Please correct the errors in the form.')
    else:
        all_shops = Shop.objects.all().order_by('name')
        default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
        form = DenominationForm()
    
    cookie_data = request.COOKIES.get('hidden_fields', '[]')
    
    try:
        # 2. Parse JSON string into a native Python list
        hidden_fields = json.loads(cookie_data)
    except json.JSONDecodeError:
        hidden_fields = []

    context = {
        'nav_title': 'Denomination',
        'form': form,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
        'hidden_fields_list': hidden_fields,
    }
    return render(request, 'entries/denomination.html', context)


@login_required
def denominations(request):
    """List all denomination groups by key with totals."""
    is_super_admin = request.user.is_superuser
    is_admin_group_user = request.user.groups.filter(name='Admin').exists()
    is_staff_group_user = request.user.groups.filter(name='Staff').exists()
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    all_shops = Shop.objects.all()
    denominations_qs = Denomination.objects.select_related('created_by')

    # Admin group users and super admins can view all; staff/others can view only their own
    if not is_admin_group_user and not is_super_admin:
        denominations_qs = denominations_qs.filter(created_by=request.user)

    denomination_groups = (
        denominations_qs.values(
            'key',
            'time_period',
            'created_by_id',
            'created_by__first_name',
            'created_by__last_name',
            'created_by__username',
            'shop__name',
            'denomination_dt'
        )
        .annotate(
            date=Min('denomination_dt'),
            total=Coalesce(
                Sum('amount'),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by('-date')
    )

    context = {
        'nav_title': 'Denomination',
        'denomination_groups': denomination_groups,
        'is_super_admin': is_super_admin,
        'is_admin': is_admin(request.user),
        'is_admin_group_user': is_admin_group_user,
        'is_staff_group_user': is_staff_group_user,
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    }
    return render(request, 'entries/denominations.html', context)


@login_required
def delete_denomination(request, key):
    """Delete denomination group by key (admin group and super admin only)."""
    if request.method != 'POST':
        return redirect('entries:denominations')

    is_super_admin = request.user.is_superuser
    is_admin_group_user = request.user.groups.filter(name='Admin').exists()
    if not is_super_admin and not is_admin_group_user:
        raise PermissionDenied('You do not have permission to delete denominations.')

    denominations_qs = Denomination.objects.filter(key=key)
    if not denominations_qs.exists():
        messages.error(request, 'Denomination not found.')
        return redirect('entries:denominations')

    deleted_count = denominations_qs.count()
    shop = denominations_qs.first().shop if denominations_qs.exists() else None
    denominations_qs.delete()

    logger.warning(
        f"Denomination group deleted by {request.user.username}: key={key}, records={deleted_count}"
    )
    log_activity(request, 'DELETE', 'Denomination', key, f'Denomination deleted: {key}', shop=shop)
    messages.success(request, 'Denomination deleted successfully!')
    return redirect('entries:denominations')


@login_required
def get_users_for_denomination(request):
    """Get all users who have created denominations"""
    
    # Admin can see all users, Staff can see only themselves
    if request.user.groups.filter(name='Admin').exists():
        users = User.objects.filter(created_denominations__isnull=False).distinct().values('id', 'username')
    else:
        # Staff and others can only see their own username
        users = User.objects.filter(id=request.user.id, created_denominations__isnull=False).distinct().values('id', 'username')
    
    return JsonResponse({'users': list(users)})


@login_required
def edit_denomination(request, key):
    """View and edit denomination by key"""
    denominations = Denomination.objects.filter(key=key).order_by('denomination')
    
    if not denominations.exists():
        messages.error(request, 'Denomination not found.')
        return redirect('entries:denominations')
    
    # Get the first record to extract metadata
    first_denom = denominations.first()
    time_period = first_denom.time_period
    created_by = first_denom.created_by
    created_at = first_denom.created_at
    
    # Authorization check: Only the creator can edit
    if created_by != request.user:
        raise PermissionDenied('You do not have permission to edit this denomination.')
    
    if request.method == 'POST':
        # Ensure time_period is in POST data (it should be from hidden input)
        post_data = request.POST.copy()
        if 'time_period' not in post_data:
            post_data['time_period'] = time_period
        
        form = DenominationForm(post_data)
        if form.is_valid():
            try:
                # Get form data including new date/time
                new_date = form.cleaned_data.get('date') or first_denom.denomination_dt
                new_time = form.cleaned_data.get('time') or timezone.localtime(timezone.now()).time()
                
                # Handle both datetime and date objects
                if hasattr(new_date, 'date'):
                    new_date = new_date.date()
                
                new_denomination_dt = timezone.make_aware(
                    datetime.combine(new_date, new_time)
                )
                
                shop = form.cleaned_data.get('shop')
                
                # Generate new key based on potentially new date
                time_period_code = {
                    'MORNING': '01',
                    'AFTERNOON': '02',
                    'EVENING': '03',
                    'NIGHT': '04'
                }.get(time_period, '00')
                new_key = f"{shop.short_name}-{new_denomination_dt.strftime('%d%m%Y')}-{time_period_code}-{request.user.username}"
                
                # Check if date has changed
                if new_key != key:
                    # Check if new date+timeperiod combination already exists
                    if Denomination.objects.filter(key=new_key).exists():
                        messages.error(request, f'Denomination for {time_period.title()} on {new_denomination_dt.strftime("%d-%m-%Y")} already exists!')
                        return render(request, 'entries/denomination.html', {
                            'form': form,
                            'key': key,
                            'time_period': time_period,
                            'shop': first_denom.shop,
                            'created_by': created_by,
                            'created_at': created_at,
                            'updated_at': denominations.order_by('-updated_at').first().updated_at,
                            'denomination_dt': first_denom.denomination_dt,
                            'is_edit_mode': True,
                            'is_super_admin': request.user.is_superuser,
                            'is_admin': is_admin(request.user),
                        })
                    
                    # Date changed, delete old records and create new ones with new key
                    old_denominations = Denomination.objects.filter(key=key)
                    old_denominations.delete()
                    key = new_key
                
                # Get form data
                note_2000 = form.cleaned_data.get('note_2000') or 0
                note_500 = form.cleaned_data.get('note_500') or 0
                note_200 = form.cleaned_data.get('note_200') or 0
                note_100 = form.cleaned_data.get('note_100') or 0
                note_50 = form.cleaned_data.get('note_50') or 0
                note_20 = form.cleaned_data.get('note_20') or 0
                note_10 = form.cleaned_data.get('note_10') or 0
                coins = form.cleaned_data.get('coins') or Decimal('0.00')
                damage = form.cleaned_data.get('damage') or Decimal('0.00')
                inside = form.cleaned_data.get('inside') or Decimal('0.00')
                bundle_500 = form.cleaned_data.get('bundle_500') or 0
                bundle_200 = form.cleaned_data.get('bundle_200') or 0
                bundle_100 = form.cleaned_data.get('bundle_100') or 0
                bundle_50 = form.cleaned_data.get('bundle_50') or 0
                bundle_20 = form.cleaned_data.get('bundle_20') or 0
                bundle_10 = form.cleaned_data.get('bundle_10') or 0
                
                # Create denomination map with new values
                denomination_updates = {
                    '2000': (note_2000, note_2000 * 2000),
                    '500': (note_500, note_500 * 500),
                    '200': (note_200, note_200 * 200),
                    '100': (note_100, note_100 * 100),
                    '50': (note_50, note_50 * 50),
                    '20': (note_20, note_20 * 20),
                    '10': (note_10, note_10 * 10),
                    'Coins': (1 if coins > 0 else 0, coins),
                    'Damage': (1 if damage > 0 else 0, damage),
                    'Inside': (1 if inside > 0 else 0, inside),
                    '500 Bundle': (bundle_500, bundle_500 * 500 * 100),
                    '200 Bundle': (bundle_200, bundle_200 * 200 * 100),
                    '100 Bundle': (bundle_100, bundle_100 * 100 * 100),
                    '50 Bundle': (bundle_50, bundle_50 * 50 * 100),
                    '20 Bundle': (bundle_20, bundle_20 * 20 * 100),
                    '10 Bundle': (bundle_10, bundle_10 * 10 * 100),
                }
                
                # Update or create denominations with new key and date
                for denom_name, (count, amount) in denomination_updates.items():
                    if count > 0 or amount > 0:
                        Denomination.objects.update_or_create(
                            key=key,
                            denomination=denom_name,
                            defaults={
                                'count': count,
                                'amount': Decimal(str(amount)),
                                'denomination_dt': new_denomination_dt,
                                'time_period': time_period,
                                'shop': shop,
                                'updated_by': request.user,
                                'created_by': created_by,
                            }
                        )
                    else:
                        Denomination.objects.filter(
                            key=key,
                            denomination=denom_name,
                        ).delete()
                
                log_activity(request, 'UPDATE', 'Denomination', key, f'Denomination updated: {key}', shop=shop)
                logger.info(f"Denomination {key} updated by {request.user.username}")
                messages.success(request, 'Denomination updated successfully!')
                return redirect('entries:view_denomination', key=key)
            except Exception as e:
                logger.error(f"Error updating denomination {key} by {request.user.username}: {str(e)}", exc_info=True)
                messages.error(request, f'An error occurred while updating denomination: {str(e)}')
        else:
            messages.error(request, f'Please correct the errors in the form: {form.errors}')
    else:
        # Pre-populate form with existing data
        initial_data = {
            'time_period': time_period,
        }
        for denom in denominations:
            if denom.denomination == '2000':
                initial_data['note_2000'] = denom.count
            elif denom.denomination == '500':
                initial_data['note_500'] = denom.count
            elif denom.denomination == '200':
                initial_data['note_200'] = denom.count
            elif denom.denomination == '100':
                initial_data['note_100'] = denom.count
            elif denom.denomination == '50':
                initial_data['note_50'] = denom.count
            elif denom.denomination == '20':
                initial_data['note_20'] = denom.count
            elif denom.denomination == '10':
                initial_data['note_10'] = denom.count
            elif denom.denomination == 'Coins':
                initial_data['coins'] = denom.amount
            elif denom.denomination == 'Damage':
                initial_data['damage'] = denom.amount
            elif denom.denomination == 'Inside':
                initial_data['inside'] = denom.amount
            elif denom.denomination == '500 Bundle':
                initial_data['bundle_500'] = denom.count
            elif denom.denomination == '200 Bundle':
                initial_data['bundle_200'] = denom.count
            elif denom.denomination == '100 Bundle':
                initial_data['bundle_100'] = denom.count
            elif denom.denomination == '50 Bundle':
                initial_data['bundle_50'] = denom.count
            elif denom.denomination == '20 Bundle':
                initial_data['bundle_20'] = denom.count
            elif denom.denomination == '10 Bundle':
                initial_data['bundle_10'] = denom.count
        
        initial_data['shop'] = first_denom.shop
        # Handle both datetime and date objects
        if hasattr(first_denom.denomination_dt, 'date'):
            initial_data['date'] = first_denom.denomination_dt.date()
            initial_data['time'] = first_denom.denomination_dt.time()
        else:
            initial_data['date'] = first_denom.denomination_dt
        form = DenominationForm(initial=initial_data)
    
    # Calculate total
    total = sum(d.amount for d in denominations)
    
    # Get the most recent updated_at timestamp
    updated_at = denominations.order_by('-updated_at').first().updated_at
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    context = {
        'nav_title': 'Denomination',
        'form': form,
        'key': key,
        'time_period': time_period,
        'shop': first_denom.shop,
        'created_by': created_by,
        'created_at': created_at,
        'updated_at': updated_at,
        'total': total,
        'denomination_dt': first_denom.denomination_dt,
        'is_edit_mode': True,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    }
    return render(request, 'entries/denomination.html', context)


@login_required
def view_denomination(request, key):
    """View denomination by key (read-only)"""
    denominations = Denomination.objects.filter(key=key).order_by('denomination')
    
    if not denominations.exists():
        messages.error(request, 'Denomination not found.')
        return redirect('entries:denominations')
    
    # Get the first record to extract metadata
    first_denom = denominations.first()
    time_period = first_denom.time_period
    created_by = first_denom.created_by
    created_at = first_denom.created_at
    
    # Authorization check: Admin can view all, Staff can view only their own
    is_admin = request.user.groups.filter(name='Admin').exists() or request.user.is_superuser
    if not is_admin and created_by != request.user:
        raise PermissionDenied('You do not have permission to view this denomination.')
    
    # Pre-populate form with existing data (read-only)
    initial_data = {
        'time_period': time_period,
    }
    for denom in denominations:
        if denom.denomination == '2000':
            initial_data['note_2000'] = denom.count
        elif denom.denomination == '500':
            initial_data['note_500'] = denom.count
        elif denom.denomination == '200':
            initial_data['note_200'] = denom.count
        elif denom.denomination == '100':
            initial_data['note_100'] = denom.count
        elif denom.denomination == '50':
            initial_data['note_50'] = denom.count
        elif denom.denomination == '20':
            initial_data['note_20'] = denom.count
        elif denom.denomination == '10':
            initial_data['note_10'] = denom.count
        elif denom.denomination == 'Coins':
            initial_data['coins'] = denom.amount
        elif denom.denomination == 'Damage':
            initial_data['damage'] = denom.amount
        elif denom.denomination == 'Inside':
            initial_data['inside'] = denom.amount
        elif denom.denomination == '500 Bundle':
            initial_data['bundle_500'] = denom.count
        elif denom.denomination == '200 Bundle':
            initial_data['bundle_200'] = denom.count
        elif denom.denomination == '100 Bundle':
            initial_data['bundle_100'] = denom.count
        elif denom.denomination == '50 Bundle':
            initial_data['bundle_50'] = denom.count
        elif denom.denomination == '20 Bundle':
            initial_data['bundle_20'] = denom.count
        elif denom.denomination == '10 Bundle':
            initial_data['bundle_10'] = denom.count
    
    initial_data['shop'] = first_denom.shop
    # Handle both datetime and date objects
    if hasattr(first_denom.denomination_dt, 'date'):
        initial_data['date'] = first_denom.denomination_dt.date()
        initial_data['time'] = first_denom.denomination_dt.time()
    else:
        initial_data['date'] = first_denom.denomination_dt
    form = DenominationForm(initial=initial_data)
    
    # Calculate total
    total = sum(d.amount for d in denominations)
    
    # Get the most recent updated_at timestamp
    updated_at = denominations.order_by('-updated_at').first().updated_at
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    context = {
        'nav_title': 'Denomination',
        'form': form,
        'key': key,
        'time_period': time_period,
        'shop': first_denom.shop,
        'total': total,
        'created_by': created_by,
        'created_at': created_at,
        'updated_at': updated_at,
        'denomination_dt': first_denom.denomination_dt,
        'is_view_mode': True,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin,
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    }
    return render(request, 'entries/denomination.html', context)


@login_required
def loan(request):
    """Create a new loan transaction"""
    if request.method != 'POST':
        return redirect('entries:add_entries')

    logger.info("============ Loan Creation Started ============")

    pawn_no    = request.POST.get('pawn_no')
    principal  = request.POST.get('principal')
    interest   = request.POST.get('interest')
    ledger_id  = request.POST.get('ledger')
    loan_type  = request.POST.get('loan_type')
    date_str   = request.POST.get('date')
    time_str   = request.POST.get('time')

    try:
        ledger           = Ledger.objects.get(id=ledger_id)
        principal_amount = Decimal(principal)
        interest_amount  = Decimal(interest)

        # ── Resolve chosen date + time ───────────────────────────────
        try:
            chosen_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.localdate()
        except ValueError:
            chosen_date = timezone.localdate()
        try:
            chosen_time = datetime.strptime(time_str, '%H:%M').time() if time_str else timezone.localtime(timezone.now()).time()
        except ValueError:
            chosen_time = timezone.localtime(timezone.now()).time()

        chosen_dt = timezone.make_aware(datetime.combine(chosen_date, chosen_time))

        logger.info(f"Loan -> type=[{loan_type}] | pawn_no=[{pawn_no}] | principal=[{principal_amount}] | interest=[{interest_amount}] | date=[{chosen_date}]")

        # ── Determine labels based on loan type ──────────────────────
        if loan_type == 'LOAN':
            name              = 'Loan'
            principal_tr_type = 'DEBIT'
            principal_remark  = 'LP'
            interest_remark   = 'LI'
            principal_rel_type = 'LOAN_PRINCIPAL'
            interest_rel_type  = 'LOAN_INTEREST'
        else:  # RELEASE
            name              = 'Release'
            principal_tr_type = 'CREDIT'
            principal_remark  = 'RP'
            interest_remark   = 'RI'
            principal_rel_type = 'RELEASE_PRINCIPAL'
            interest_rel_type  = 'RELEASE_INTEREST'

        # ── Resolve linked accounts from BT_Ledger_Accounts ─────────
        principal_account = transaction_helper.get_linked_account(ledger, principal_rel_type)
        interest_account  = transaction_helper.get_linked_account(ledger, interest_rel_type)
        logger.info(f"Linked accounts -> principal=[{principal_account}] | interest=[{interest_account}]")

        if principal_account is None or interest_account is None:
            messages.error(request, 'No linked accounts found for principal and interest. Please check ledger configuration.')
            return redirect('entries:add_entries')
        
        with db_transaction.atomic():
            shop = Shop.objects.select_for_update().get(pk=ledger.shop_id)

            # ── Step 1: Balance check for LOAN type ──────────────────
            if loan_type == 'LOAN':
                available = transaction_helper.get_balance(shop, chosen_date)
                logger.info(f"Balance check -> shop=[{shop.short_name}] | available=[{available}] | required=[{principal_amount}]")
                if principal_amount > available:
                    messages.error(request, f'Insufficient balance in {shop.short_name}. Available: {available}')
                    return redirect('entries:add_entries')

            # ── Step 2: Create loan record ───────────────────────────
            loan_entry = Loan(
                pawn_no=pawn_no,
                shop=shop,
                ledger=ledger,
                type=loan_type,
                principal=principal_amount,
                interest=interest_amount,
                transaction_dt=chosen_dt,
                created_by=request.user,
                updated_by=request.user
            )
            loan_entry.save()
            logger.info(f"Loan record created -> id=[{loan_entry.id}]")

            # ── Step 3: Find existing principal/interest transactions for the day ──
            logger.info(f"Searching transactions -> shop=[{shop.short_name}] | date=[{chosen_date}]")
            principal_filter = dict(shop=shop, loan_tr_type__startswith=principal_remark, transaction_dt__date=chosen_date)
            interest_filter  = dict(shop=shop, loan_tr_type__startswith=interest_remark,  transaction_dt__date=chosen_date)
            if principal_account:
                principal_filter['acc'] = principal_account
            if interest_account:
                interest_filter['acc'] = interest_account

            existing_principal = Transactions.objects.filter(**principal_filter).first()
            existing_interest  = Transactions.objects.filter(**interest_filter).first()
            logger.info(f"Existing principal found: [{existing_principal is not None}] | Existing interest found: [{existing_interest is not None}]")

            # ── Step 4: Update or create principal transaction ───────
            if principal_amount > 0:
                transaction_helper._add_or_create_transaction(
                    trans=existing_principal,
                    amount=principal_amount,
                    loan_tr_type=principal_remark,
                    remark="",
                    tr_type=principal_tr_type,
                    shop=shop,
                    chosen_dt=chosen_dt,
                    user=request.user,
                    account=principal_account,
                    label="principal"
                )
            else:
                logger.info("[principal] amount is 0 — skipping transaction")

            # ── Step 5: Update or create interest transaction ────────
            if interest_amount > 0:
                transaction_helper._add_or_create_transaction(
                    trans=existing_interest,
                    amount=interest_amount,
                    remark="",
                    loan_tr_type=interest_remark,
                    tr_type='CREDIT',
                    shop=shop,
                    chosen_dt=chosen_dt + timedelta(milliseconds=10),
                    user=request.user,
                    account=interest_account,
                    label="interest"
                )
            else:
                logger.info("[interest] amount is 0 — skipping transaction")

        # ── Step 6: Append pawn_no to transaction remarks ────────
        # Re-fetch transactions after create/update to get latest state
        updated_principal_trans = Transactions.objects.filter(**principal_filter).first()
        # updated_interest_trans  = Transactions.objects.filter(**interest_filter).first()

        if updated_principal_trans:
            transaction_helper.append_pawn_no_to_remark(
                updated_principal_trans, pawn_no, request.user, label="principal remark"
            )

        log_activity(request, 'CREATE', 'Loan', loan_entry.id, f'Loan created: {loan_entry.pawn_no}', shop=shop)
        messages.success(request, f'{loan_type.capitalize()} entry created successfully!')
        logger.info(f"Loan [{loan_entry.id}] created by [{request.user.username}]")
        logger.info("============ Loan Creation Completed ============")

    except Ledger.DoesNotExist:
        logger.error(f"Ledger [{ledger_id}] not found")
        messages.error(request, 'Selected ledger does not exist.')
    except Exception as e:
        logger.error(f"Error creating loan: {str(e)}", exc_info=True)
        messages.error(request, f'Error creating loan entry: {str(e)}')

    return redirect('entries:add_entries')

@login_required
def bulk_loan(request):
    """Create multiple loan/release transactions in one atomic operation."""
    if request.method != 'POST':
        return redirect('entries:add_entries')

    logger.info("============ Bulk Loan Creation Started ============")

    loan_type   = request.POST.get('loan_type')
    date_str    = request.POST.get('date')
    time_str    = request.POST.get('time')
    pawn_nos    = request.POST.getlist('pawn_no[]')
    ledger_ids  = request.POST.getlist('ledger[]')
    principals  = request.POST.getlist('principal[]')
    interests   = request.POST.getlist('interest[]')

    if not pawn_nos:
        messages.error(request, 'No loan entries submitted.')
        return redirect('entries:add_entries')

    # ── Resolve date + time ──────────────────────────────────────────
    try:
        chosen_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.localdate()
    except ValueError:
        chosen_date = timezone.localdate()
    try:
        chosen_time = datetime.strptime(time_str, '%H:%M').time() if time_str else timezone.localtime(timezone.now()).time()
    except ValueError:
        chosen_time = timezone.localtime(timezone.now()).time()

    chosen_dt = timezone.make_aware(datetime.combine(chosen_date, chosen_time))

    # ── Determine type labels ────────────────────────────────────────
    if loan_type == 'LOAN':
        principal_tr_type  = 'DEBIT'
        principal_loan_type = 'LP'
        interest_loan_type  = 'LI'
        principal_rel_type  = 'LOAN_PRINCIPAL'
        interest_rel_type   = 'LOAN_INTEREST'
    else:
        principal_tr_type  = 'CREDIT'
        principal_loan_type = 'RP'
        interest_loan_type  = 'RI'
        principal_rel_type  = 'RELEASE_PRINCIPAL'
        interest_rel_type   = 'RELEASE_INTEREST'

    created_loans = []
    t_user = User.objects.filter(username='system').first()
    try:
        with db_transaction.atomic():
            for idx, (pawn_no, ledger_id, principal, interest) in enumerate(
                zip(pawn_nos, ledger_ids, principals, interests), start=1
            ):
                logger.info(f"[Row {idx}] pawn_no=[{pawn_no}] | ledger=[{ledger_id}] | principal=[{principal}] | interest=[{interest}]")

                try:
                    ledger           = Ledger.objects.get(id=ledger_id)
                    principal_amount = Decimal(principal)
                    interest_amount  = Decimal(interest)
                except (Ledger.DoesNotExist, ValueError) as e:
                    raise Exception(f'Row {idx} ({pawn_no}): {str(e)}')

                shop = Shop.objects.select_for_update().get(pk=ledger.shop_id)

                principal_account = transaction_helper.get_linked_account(ledger, principal_rel_type)
                interest_account  = transaction_helper.get_linked_account(ledger, interest_rel_type)

                if principal_account is None or interest_account is None:
                    raise Exception(f'Row {idx} ({pawn_no}): No linked accounts found for ledger [{ledger.name}].')

                # ── Balance check for LOAN ───────────────────────────
                if loan_type == 'LOAN':
                    available = transaction_helper.get_balance(shop, chosen_date)
                    if principal_amount > available:
                        raise Exception(f'Row {idx} ({pawn_no}): Insufficient balance in {shop.short_name}. Available: {available}')

                # ── Create loan record ───────────────────────────────
                loan_entry = Loan(
                    pawn_no=pawn_no,
                    shop=shop,
                    ledger=ledger,
                    type=loan_type,
                    principal=principal_amount,
                    interest=interest_amount,
                    transaction_dt=chosen_dt + timedelta(milliseconds=idx),
                    created_by=request.user,
                    updated_by=request.user,
                )
                loan_entry.save()
                logger.info(f"[Row {idx}] Loan record created -> id=[{loan_entry.id}]")

                # ── Find or create principal transaction ─────────────
                principal_filter = dict(
                    shop=shop,
                    loan_tr_type=principal_loan_type,
                    transaction_dt__date=chosen_date,
                    acc=principal_account,
                )
                interest_filter = dict(
                    shop=shop,
                    loan_tr_type=interest_loan_type,
                    transaction_dt__date=chosen_date,
                    acc=interest_account,
                )

                existing_principal = Transactions.objects.filter(**principal_filter).first()
                existing_interest  = Transactions.objects.filter(**interest_filter).first()

                if principal_amount > 0:
                    transaction_helper._add_or_create_transaction(
                        trans=existing_principal,
                        amount=principal_amount,
                        loan_tr_type=principal_loan_type,
                        remark="",
                        tr_type=principal_tr_type,
                        shop=shop,
                        chosen_dt=chosen_dt,
                        user=t_user,
                        account=principal_account,
                        label=f"row{idx} principal",
                    )

                if interest_amount > 0:
                    transaction_helper._add_or_create_transaction(
                        trans=existing_interest,
                        amount=interest_amount,
                        loan_tr_type=interest_loan_type,
                        remark="",
                        tr_type='CREDIT',
                        shop=shop,
                        chosen_dt=chosen_dt + timedelta(milliseconds=10),
                        user=t_user,
                        account=interest_account,
                        label=f"row{idx} interest",
                    )

                created_loans.append((loan_entry, pawn_no, principal_filter))

            # ── After all rows saved: update principal remarks ───────
            for loan_entry, pawn_no, principal_filter in created_loans:
                updated_principal = Transactions.objects.filter(**principal_filter).first()
                if updated_principal:
                    mode = "range" if loan_type == "LOAN" else "list"
                    transaction_helper.append_pawn_no_to_remark(
                        updated_principal, pawn_no, request.user,
                        label=f"principal remark [{pawn_no}]",
                    )

            for loan_entry, pawn_no, _ in created_loans:
                log_activity(
                    request, 'CREATE', 'Loan', loan_entry.id,
                    f'{loan_type.capitalize()} created: {pawn_no}', shop=loan_entry.shop
                )

        count = len(created_loans)
        messages.success(request, f'{count} {loan_type.capitalize()} entr{"y" if count == 1 else "ies"} created successfully!')
        logger.info(f"Bulk loan complete -> count=[{count}] | type=[{loan_type}] | user=[{request.user.username}]")
        logger.info("============ Bulk Loan Creation Completed ============")

    except Exception as e:
        logger.error(f"Bulk loan error: {str(e)}", exc_info=True)
        messages.error(request, f'Error: {str(e)}')

    return redirect('entries:add_entries')

@login_required
@ensure_csrf_cookie
def loans(request):
    """View all loan transactions with pagination and filtering"""
    loans_list = Loan.objects.select_related('ledger', 'created_by', 'updated_by').order_by('-transaction_dt')
    
    # Apply filters
    from_date = (request.GET.get('from_date') or '').strip()
    to_date = (request.GET.get('to_date') or '').strip()
    ledger_filter = (request.GET.get('ledger') or '').strip()
    shop_filter = (request.GET.get('shop') or '').strip()
    type_filter = (request.GET.get('type') or '').strip().upper()
    search_query = (request.GET.get('search') or '').strip()

    query_params = request.GET.copy()
    query_params.pop('page', None)
    filter_query = query_params.urlencode()

    if not shop_filter:
        default_shop_short_name = Configuration.get_value(Configuration.Key.DEFAULT_SHOP, default='')
        if default_shop_short_name:
            default_shop = Shop.objects.filter(short_name=default_shop_short_name).first()
            if default_shop:
                shop_filter = str(default_shop.id)

    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            loans_list = loans_list.filter(transaction_dt__date__gte=from_date_obj)
        except ValueError:
            from_date = ''

    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            loans_list = loans_list.filter(transaction_dt__date__lte=to_date_obj)
        except ValueError:
            to_date = ''
    
    if shop_filter != 'all':
        print(f"Filtering by shop: {shop_filter}")
        loans_list = loans_list.filter(shop_id=shop_filter)

    if ledger_filter:
        loans_list = loans_list.filter(ledger_id=ledger_filter)
    elif ledger_filter:
        ledger_filter = ''

    if type_filter in ['LOAN', 'RELEASE']:
        loans_list = loans_list.filter(type=type_filter)
    else:
        type_filter = ''

    if search_query:
        loans_list = loans_list.filter(pawn_no__icontains=search_query)
    
    # Calculate totals
    loan_totals = loans_list.filter(type='LOAN').aggregate(
        total_principal=Sum('principal'),
        total_interest=Sum('interest')
    )
    release_totals = loans_list.filter(type='RELEASE').aggregate(
        total_principal=Sum('principal'),
        total_interest=Sum('interest')
    )
    
    # Pagination
    paginator = Paginator(loans_list, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all ledgers for filter dropdown
    all_ledgers = Ledger.objects.all().order_by('name')
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    
    context = {
        'nav_title': 'Loan Transactions',
        'page_obj': page_obj,
        'all_ledgers': all_ledgers,
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
        'from_date': from_date,
        'to_date': to_date,
        'ledger_filter': ledger_filter,
        'shop_filter': shop_filter,
        'type_filter': type_filter,
        'search_query': search_query,
        'loan_totals': loan_totals,
        'release_totals': release_totals,
        'all_loans': loans_list,  # For print view
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'filter_query': filter_query,
        'is_admin_user': is_admin(request.user) or request.user.is_superuser,
    }
    return render(request, 'entries/loans.html', context)


@login_required
def edit_loan(request, pk):
    """Edit a loan transaction"""
    loan = get_object_or_404(Loan, pk=pk)

    # ── Snapshot old values before any changes ──────────────────────
    old_principal    = loan.principal
    old_interest     = loan.interest
    old_type         = loan.type
    old_ledger       = loan.ledger
    old_shop         = old_ledger.shop
    old_date         = timezone.localtime(loan.transaction_dt).date()
    old_pawn_no      = loan.pawn_no
    t_user = User.objects.filter(username='system').first()

    if request.method == 'POST':
        logger.info("============ Loan Update Started ============")
        form = LoanEditForm(request.POST, instance=loan)

        if form.is_valid():
            updated_loan     = form.save(commit=False)
            new_principal    = updated_loan.principal
            new_interest     = updated_loan.interest
            new_type         = updated_loan.type
            new_ledger       = updated_loan.ledger
            new_shop         = new_ledger.shop

            # Resolve new transaction date
            chosen_date = form.cleaned_data.get('date')
            chosen_time = form.cleaned_data.get('time') or timezone.localtime(timezone.now()).time()
            new_dt      = timezone.make_aware(datetime.combine(chosen_date, chosen_time))
            new_date    = chosen_date

            logger.info(f"Old -> type=[{old_type}] | principal=[{old_principal}] | interest=[{old_interest}] | shop=[{old_shop}] | date=[{old_date}]")
            logger.info(f"New -> type=[{new_type}] | principal=[{new_principal}] | interest=[{new_interest}] | shop=[{new_shop}] | date=[{new_date}]")

            # Remark labels
            old_principal_remark = f"{'L' if old_type == 'LOAN' else 'R'}P"
            old_interest_remark  = f"{'L' if old_type == 'LOAN' else 'R'}I"
            new_principal_remark = f"{'L' if new_type == 'LOAN' else 'R'}P"
            new_interest_remark  = f"{'L' if new_type == 'LOAN' else 'R'}I"

            # Resolve linked accounts for old and new types
            old_principal_rel = 'LOAN_PRINCIPAL'   if old_type == 'LOAN' else 'RELEASE_PRINCIPAL'
            old_interest_rel  = 'LOAN_INTEREST'    if old_type == 'LOAN' else 'RELEASE_INTEREST'
            new_principal_rel = 'LOAN_PRINCIPAL'   if new_type == 'LOAN' else 'RELEASE_PRINCIPAL'
            new_interest_rel  = 'LOAN_INTEREST'    if new_type == 'LOAN' else 'RELEASE_INTEREST'

            old_principal_account = transaction_helper.get_linked_account(old_ledger, old_principal_rel)
            old_interest_account  = transaction_helper.get_linked_account(old_ledger, old_interest_rel)
            new_principal_account = transaction_helper.get_linked_account(new_ledger, new_principal_rel)
            new_interest_account  = transaction_helper.get_linked_account(new_ledger, new_interest_rel)
            logger.info(
                f"Linked accounts -> old_principal=[{old_principal_account}] | old_interest=[{old_interest_account}] | "
                f"new_principal=[{new_principal_account}] | new_interest=[{new_interest_account}]"
            )

            same_shop = (old_shop.id == new_shop.id)
            same_date = (old_date == new_date)
            type_changed = (old_type != new_type)

            try:
                with db_transaction.atomic():

                    # ── Step 1: Validate balance for LOAN type ──────────────
                    if new_type == 'LOAN':
                        available = transaction_helper.get_balance(new_shop, new_date)
                        # If same shop and old type was LOAN, add back old principal
                        # because it will be reversed before new amount is applied
                        if same_shop and old_type == 'LOAN':
                            available += old_principal
                        logger.info(f"Balance check -> available=[{available}] | required=[{new_principal}]")
                        if new_principal > available:
                            form.add_error(None, f'Insufficient balance in {new_shop.short_name} on {new_date}. Available: {available}')
                            return render(request, 'entries/edit-loan.html', {
                                'nav_title': 'Loan Transactions',
                                'form': form,
                                'loan': loan,
                            })

                    # ── Step 2: Find old transactions on old date ───────────
                    logger.info(f"Searching old transactions -> shop=[{old_shop.short_name}] | date=[{old_date}]")
                    old_principal_filter = dict(shop=old_shop, loan_tr_type__startswith=old_principal_remark, transaction_dt__date=old_date)
                    old_interest_filter  = dict(shop=old_shop, loan_tr_type__startswith=old_interest_remark,  transaction_dt__date=old_date)
                    if old_principal_account:
                        old_principal_filter['acc'] = old_principal_account
                    if old_interest_account:
                        old_interest_filter['acc'] = old_interest_account

                    old_principal_trans = Transactions.objects.filter(**old_principal_filter).first()
                    old_interest_trans  = Transactions.objects.filter(**old_interest_filter).first()
                    logger.info(f"Old principal trans found: [{old_principal_trans is not None}] | Old interest trans found: [{old_interest_trans is not None}]")

                    # ── Step 3: Find new transactions on new date ───────────
                    logger.info(f"Searching new transactions -> shop=[{new_shop.short_name}] | date=[{new_date}]")
                    new_principal_filter = dict(shop=new_shop, loan_tr_type__startswith=new_principal_remark, transaction_dt__date=new_date)
                    new_interest_filter  = dict(shop=new_shop, loan_tr_type__startswith=new_interest_remark,  transaction_dt__date=new_date)
                    if new_principal_account:
                        new_principal_filter['acc'] = new_principal_account
                    if new_interest_account:
                        new_interest_filter['acc'] = new_interest_account

                    new_principal_trans = Transactions.objects.filter(**new_principal_filter).first()
                    new_interest_trans  = Transactions.objects.filter(**new_interest_filter).first()
                    logger.info(f"New principal trans found: [{new_principal_trans is not None}] | New interest trans found: [{new_interest_trans is not None}]")

                    

                    # ── Step 4: Reduce old amounts from old transactions ─────
                    # Only needed when shop or date changed
                    if not (same_shop and same_date):
                        logger.info("Shop or date changed — reversing old transactions")
                        transaction_helper._reduce_or_delete_transaction(old_principal_trans, old_principal, t_user, label="old principal")
                        transaction_helper._reduce_or_delete_transaction(old_interest_trans,  old_interest,  t_user, label="old interest")

                    # ── Step 5: Apply new amounts to new transactions ────────
                    if same_shop and same_date:
                        logger.info("Same shop & date — checking for type change")

                        if type_changed:
                            # Type changed — delete old transactions and create new ones
                            logger.info(f"Loan type changed [{old_type}] -> [{new_type}] — removing old, creating new")

                            transaction_helper._reduce_or_delete_transaction(old_principal_trans, old_principal, t_user, label="old principal (type change)")
                            transaction_helper._reduce_or_delete_transaction(old_interest_trans,  old_interest,  t_user, label="old interest (type change)")

                            transaction_helper._add_or_create_transaction(
                                trans=new_principal_trans,
                                amount=new_principal,
                                remark="",
                                tr_type='DEBIT' if new_type == 'LOAN' else 'CREDIT',
                                shop=new_shop,
                                chosen_dt=new_dt,
                                user=t_user,
                                account=new_principal_account,
                                loan_tr_type='LP' if new_type == 'LOAN' else 'RP',
                                label="new principal (type change)"
                            )
                            transaction_helper._add_or_create_transaction(
                                trans=new_interest_trans,
                                amount=new_interest,
                                remark="",
                                tr_type='CREDIT',
                                shop=new_shop,
                                chosen_dt=new_dt,
                                user=t_user,
                                account=new_interest_account,
                                loan_tr_type='LI' if new_type == 'LOAN' else 'RI',
                                label="new interest (type change)"
                            )
                        else:
                            # Same type — just update amounts by delta
                            logger.info("Same type — updating amounts in place")
                            transaction_helper._apply_amount_delta(
                                trans=old_principal_trans,
                                old_amount=old_principal,
                                new_amount=new_principal,
                                remark="",
                                tr_type='DEBIT' if new_type == 'LOAN' else 'CREDIT',
                                shop=new_shop,
                                chosen_dt=new_dt,
                                user=t_user,
                                account=new_principal_account,
                                loan_tr_type='LP' if new_type == 'LOAN' else 'RP',
                                label="principal"
                            )
                            transaction_helper._apply_amount_delta(
                                trans=old_interest_trans,
                                old_amount=old_interest,
                                new_amount=new_interest,
                                remark="",
                                tr_type='CREDIT',
                                shop=new_shop,
                                chosen_dt=new_dt,
                                user=t_user,
                                account=new_interest_account,
                                loan_tr_type='LI' if new_type == 'LOAN' else 'RI',
                                label="interest"
                            )
                    else:
                        # Different shop or date — add to new transactions
                        logger.info("Different shop or date — applying to new transactions")
                        transaction_helper._add_or_create_transaction(
                            trans=new_principal_trans,
                            amount=new_principal,
                            remark="",
                            tr_type='DEBIT' if new_type == 'LOAN' else 'CREDIT',
                            shop=new_shop,
                            chosen_dt=new_dt,
                            user=t_user,
                            account=new_principal_account,
                            loan_tr_type='LP' if new_type == 'LOAN' else 'RP',
                            label="new principal"
                        )
                        transaction_helper._add_or_create_transaction(
                            trans=new_interest_trans,
                            amount=new_interest,
                            remark="",
                            tr_type='CREDIT',
                            shop=new_shop,
                            chosen_dt=new_dt,
                            user=t_user,
                            account=new_interest_account,
                            loan_tr_type='LI' if new_type == 'LOAN' else 'RI',
                            label="new interest"
                        )

                    # ── Step 6: Save updated loan ───────────────────────────
                    # Always remove old pawn_no first, then append new one below
                    old_p = Transactions.objects.filter(**old_principal_filter).first()
                    if old_p:
                        transaction_helper.remove_pawn_no_from_remark(
                            old_p, old_pawn_no, request.user, label="old principal remark"
                        )

                    # Add pawn_no to new/updated transactions
                    new_p = Transactions.objects.filter(**new_principal_filter).first()
                    if new_p:
                        transaction_helper.append_pawn_no_to_remark(
                            new_p, updated_loan.pawn_no, request.user, label="new principal remark"
                        )
                    
                    updated_loan.shop = new_shop
                    updated_loan.updated_by = request.user
                    updated_loan.transaction_dt = new_dt
                    updated_loan.save()
                    logger.info(f"Loan [{pk}] saved -> type=[{new_type}] | principal=[{new_principal}] | interest=[{new_interest}]")

                log_activity(request, 'UPDATE', 'Loan', updated_loan.id, f'Loan updated: {updated_loan.pawn_no}', shop=updated_loan.shop)
                messages.success(request, 'Loan transaction updated successfully!')
                logger.info("============ Loan Update Completed ============")
                return redirect('entries:loans')

            except Exception as e:
                logger.error(f"Error updating loan [{pk}]: {str(e)}", exc_info=True)
                messages.error(request, f'An error occurred while updating loan: {str(e)}')

        else:
            messages.error(request, 'Please correct the errors below.')

    else:
        form = LoanEditForm(
            instance=loan,
            initial={
                'date': timezone.localtime(loan.transaction_dt).date(),
                'time': timezone.localtime(loan.transaction_dt).time().replace(second=0, microsecond=0),
            }
        )

    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    all_shops = Shop.objects.all()
    return render(request, 'entries/edit-loan.html', {
        'nav_title': 'Loan Transactions',
        'form': form,
        'loan': loan,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    })


@login_required
@csrf_protect
@admin_required
def delete_loan(request, pk):
    """Delete a loan transaction (admin only)"""
    loan = get_object_or_404(Loan, pk=pk)

    if request.method == 'POST':
        loan_id        = loan.id
        loan_pawn_no   = loan.pawn_no
        loan_type      = loan.type
        loan_shop      = loan.ledger.shop
        loan_date      = timezone.localtime(loan.transaction_dt).date()
        loan_principal = loan.principal
        loan_interest  = loan.interest
        t_user = User.objects.filter(username='system').first()

        principal_remark = 'LP' if loan_type == 'LOAN' else 'RP'
        interest_remark  = 'LI' if loan_type == 'LOAN' else 'RI'

        logger.info("============ Loan Deletion Started ============")
        logger.info(f"Loan -> id=[{loan_id}] | pawn_no=[{loan_pawn_no}] | type=[{loan_type}] | shop=[{loan_shop.short_name}] | date=[{loan_date}]")
        logger.info(f"Amounts -> principal=[{loan_principal}] | interest=[{loan_interest}]")

        try:
            with db_transaction.atomic():

                # ── Step 1: Find principal and interest transactions ─────
                logger.info(f"Searching transactions -> shop=[{loan_shop.short_name}] | date=[{loan_date}]")

                principal_rel_type = 'LOAN_PRINCIPAL' if loan_type == 'LOAN' else 'RELEASE_PRINCIPAL'
                interest_rel_type  = 'LOAN_INTEREST'  if loan_type == 'LOAN' else 'RELEASE_INTEREST'
                principal_account = transaction_helper.get_linked_account(loan.ledger, principal_rel_type)
                interest_account  = transaction_helper.get_linked_account(loan.ledger, interest_rel_type)

                principal_filter = dict(shop=loan_shop, loan_tr_type__startswith=principal_remark, transaction_dt__date=loan_date)
                interest_filter  = dict(shop=loan_shop, loan_tr_type__startswith=interest_remark,  transaction_dt__date=loan_date)
                if principal_account:
                    principal_filter['acc'] = principal_account
                if interest_account:
                    interest_filter['acc'] = interest_account

                principal_trans = Transactions.objects.filter(**principal_filter).first()
                interest_trans  = Transactions.objects.filter(**interest_filter).first()
                logger.info(f"Principal trans found: [{principal_trans is not None}] | Interest trans found: [{interest_trans is not None}]")

                # ── Step 1b: Remove pawn_no from remarks before reduction ──
                if principal_trans:
                    transaction_helper.remove_pawn_no_from_remark(
                        principal_trans, loan_pawn_no, request.user, label="principal remark"
                    )
                if interest_trans:
                    transaction_helper.remove_pawn_no_from_remark(
                        interest_trans, loan_pawn_no, request.user, label="interest remark"
                    )

                # ── Step 2: Reduce or delete principal transaction ───────
                transaction_helper._reduce_or_delete_transaction(
                    trans=principal_trans,
                    amount=loan_principal,
                    user=t_user,
                    label="principal"
                )

                # ── Step 3: Reduce or delete interest transaction ────────
                transaction_helper._reduce_or_delete_transaction(
                    trans=interest_trans,
                    amount=loan_interest,
                    user=t_user,
                    label="interest"
                )

                # ── Step 4: Delete the loan record ───────────────────────
                loan.delete()
                logger.warning(f"Loan deleted by [{request.user.username}] -> id=[{loan_id}] | type=[{loan_type}] | pawn_no=[{loan_pawn_no}]")

            log_activity(request, 'DELETE', 'Loan', loan.id, f'Loan deleted: {loan.pawn_no}', shop=loan.shop)
            messages.success(request, 'Loan transaction deleted successfully!')
            logger.info("============ Loan Deletion Completed ============")

        except Exception as e:
            logger.error(f"Error deleting loan [{loan_id}] by [{request.user.username}]: {str(e)}", exc_info=True)
            messages.error(request, 'An error occurred while deleting loan.')

        return redirect('entries:loans')

    return render(request, 'entries/delete-loan.html', {
        'nav_title': 'Loan Transactions',
        'loan': loan,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
    })


@login_required
@admin_required
def transaction_history(request, pk):
    """Show current transaction and full audit history for a transaction record."""
    transaction = get_object_or_404(
        Transactions.objects.select_related('shop', 'created_by', 'updated_by'),
        pk=pk
    )
    # Exclude creation record to avoid duplication with current record
    history_records = transaction.history.all().order_by('-history_date') 
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    context = {
        'nav_title': 'Other Transactions',
        'transaction': transaction,
        'history_records': history_records,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    }
    return render(request, 'entries/transaction_history.html', context)


@login_required
@admin_required
def loan_history(request, pk):
    """Show current loan entry and full audit history for a loan record."""
    loan = get_object_or_404(
        Loan.objects.select_related('ledger', 'created_by', 'updated_by'),
        pk=pk
    )
    history_records = loan.history.all().order_by('-history_date')
    all_shops = Shop.objects.all().order_by('name')
    default_shop_short_name = Configuration.objects.filter(key=Configuration.Key.DEFAULT_SHOP).first()
    
    context = {
        'nav_title': 'Loan Transactions',
        'loan': loan,
        'history_records': history_records,
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
        'all_shops': all_shops,
        'default_shop_short_name': default_shop_short_name,
    }
    return render(request, 'entries/loan_history.html', context)


def custom_404_view(request, exception=None):
    """Custom 404 error handler that properly passes request context"""
    return render(request, '404.html', status=404)


def custom_403_view(request, exception=None):
    """Custom 403 error handler for permission denied"""
    return render(request, '403.html', status=403)

def about(request):
    """About page with system information and credits"""
    context = {
        'nav_title': 'About',
        'is_super_admin': request.user.is_superuser,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'entries/about.html', context)

def document_view(request, filename=None):
    docs_dir = Path("docs")

    # File list is always needed (for the sidebar)
    files = sorted(f.stem for f in docs_dir.glob("*.md"))

    content = None
    title = None

    if filename is None:
        filename = files[0]
    file_path = docs_dir / f"{filename}.md"
    if not file_path.exists():
        return render(request, "404.html", status=404)

    with open(file_path, "r", encoding="utf-8") as f:
        md_content = f.read()

    content = markdown.markdown(
        md_content,
        extensions=["fenced_code", "tables", "toc"],
    )
    title = filename  # matched against `file` in sidebar loop

    return render(
        request,
        "entries/doc-view.html",      # ← single template now
        {
            "files": files,
            "content": content,
            "title": title,
            'nav_title': 'Help',
            'is_super_admin': request.user.is_superuser,
            'is_admin': is_admin(request.user),
        },
    )