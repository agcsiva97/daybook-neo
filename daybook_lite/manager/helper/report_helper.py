"""
Report generation helper methods for PDF, CSV, and Excel formats.
Uses WeasyPrint for PDF generation with support for multiple fonts.
"""

import io
import csv
from datetime import datetime
from decimal import Decimal
import os

from openpyxl import Workbook
from django.conf import settings
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_pdf_report(title, table_data):
    """
    Generate a PDF report using WeasyPrint with the given title and table data.
    A4 page size, supports multiple fonts including Tamil and Google Sans.
    
    Args:
        title (str): The title of the report
        table_data (list): List of lists representing table rows
                          First row should be headers
                          e.g., [['Header1', 'Header2'], ['Value1', 'Value2'], ...]
    
    Returns:
        io.BytesIO: PDF file content as bytes
    """
    # Lazy import to avoid module-level import errors on Windows without GTK
    try:
        from weasyprint import HTML
    except ImportError as e:
        raise ImportError(
            f"WeasyPrint is required for PDF generation. "
            f"Please install it with: pip install weasyprint\n"
            f"For Windows, please follow: https://doc.courtbouillon.org/weasyprint/stable/first_steps.html#installation\n"
            f"Error: {e}"
        )
    
    # Get font paths
    static_dir = settings.STATIC_ROOT if not settings.DEBUG else settings.STATICFILES_DIRS[0]
    google_sans_path = os.path.join(
        static_dir, 'webfonts', 'Google_Sans', 
        'GoogleSans-VariableFont_GRAD,opsz,wght.ttf'
    )
    
    # Format table rows into HTML
    table_rows_html = ""
    
    if table_data:
        for i, row in enumerate(table_data):
            row_html = ""
            for cell in row:
                if isinstance(cell, Decimal):
                    value = f"{cell:.2f}"
                else:
                    value = str(cell) if cell is not None else ""
                
                # Try to determine if numeric for alignment
                is_numeric = False
                try:
                    float(value.replace(',', ''))
                    is_numeric = True
                except (ValueError, AttributeError):
                    pass
                
                align = "right" if is_numeric else "left"
                row_html += f'<td style="text-align:{align}; padding: 8px; border: 1px solid #ddd;">{value}</td>'
            
            # Header row styling
            if i == 0:
                table_rows_html += f"<tr class='header'>{row_html}</tr>"
            else:
                # Alternating row colors for readability
                bg_color = "#ffffff" if i % 2 == 1 else "#f9f9f9"
                table_rows_html += f"<tr style='background-color: {bg_color};'>{row_html}</tr>"
    
    # HTML template with embedded fonts and A4 page sizing
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @page {{
                size: A4;
                margin: 20mm;
            }}
            
            @font-face {{
                font-family: 'GoogleSans';
                src: url('file:///{google_sans_path}') format('truetype');
            }}
            
            * {{
                font-family: 'GoogleSans', Arial, sans-serif;
            }}
            
            body {{
                margin: 0;
                padding: 0;
                color: #333;
                font-size: 12px;
            }}
            
            h1 {{
                text-align: center;
                color: #1f4788;
                margin: 0 0 10px 0;
                font-size: 18px;
            }}
            
            .timestamp {{
                text-align: right;
                font-size: 10px;
                color: #999;
                margin-bottom: 15px;
            }}
            
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
            }}
            
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                font-size: 11px;
            }}
            
            tr.header {{
                background-color: #1f4788;
                color: white;
                font-weight: bold;
                text-align: center;
            }}
            
            tr.header td {{
                color: white;
                border-color: #1f4788;
                padding: 10px 8px;
            }}
        </style>
    </head>
    <body>
        <h1>{title}</h1>
        <div class="timestamp">Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}</div>
        <table>
            {table_rows_html}
        </table>
    </body>
    </html>
    """
    
    # Generate PDF
    try:
        pdf_file = io.BytesIO()
        HTML(string=html_content, base_url=".").write_pdf(pdf_file)
        pdf_file.seek(0)
        return pdf_file
    except Exception as e:
        raise Exception(f"Error generating PDF: {e}")


def generate_csv_report(title, table_data):
    """
    Generate a CSV report with the given title and table data.
    
    Args:
        title (str): The title of the report
        table_data (list): List of lists representing table rows
                          First row should be headers
    
    Returns:
        io.BytesIO: CSV file content as bytes
    """
    csv_buffer = io.BytesIO()
    # Use StringIO wrapper for text mode
    text_buffer = io.StringIO()
    
    writer = csv.writer(text_buffer)
    
    # Write title as first row
    if title:
        writer.writerow([title])
        writer.writerow([])  # Empty row for spacing
    
    # Write timestamp
    writer.writerow(['Generated on:', datetime.now().strftime('%d-%m-%Y %H:%M:%S')])
    writer.writerow([])  # Empty row for spacing
    
    # Write table data
    if table_data:
        for row in table_data:
            formatted_row = []
            for cell in row:
                if isinstance(cell, Decimal):
                    formatted_row.append(str(cell))
                else:
                    formatted_row.append(str(cell) if cell is not None else '')
            writer.writerow(formatted_row)
    
    # Convert StringIO to BytesIO
    csv_buffer = io.BytesIO(text_buffer.getvalue().encode('utf-8'))
    csv_buffer.seek(0)
    return csv_buffer


def build_trial_balance_table(rows, total_debit, total_credit):
    """
    Build table_data list for trial balance reports.

    Args:
        rows (list): Dicts with keys 'name', 'debit', 'credit'.
        total_debit (Decimal): Column total for debits.
        total_credit (Decimal): Column total for credits.

    Returns:
        list[list]: table_data suitable for generate_excel_report / generate_csv_report.
    """
    table = [['Account Type', 'Debit', 'Credit']]
    for row in rows:
        table.append([row['name'], row['debit'] if row['debit'] != '' else '', row['credit'] if row['credit'] != '' else ''])
    table.append(['Total', total_debit, total_credit])
    return table


def build_bs_shop_table(group_fy_data, total_opening, total_closing):
    """
    Build table_data list for the Balance Sheet (with accounts level).

    Args:
        group_fy_data (list): Processed group data with nested types and accounts.
        total_opening (Decimal): Grand total opening.
        total_closing (Decimal): Grand total closing.

    Returns:
        list[list]: table_data suitable for generate_excel_report / generate_csv_report.
    """
    table = [['Group / Type / Account', 'Opening', 'Closing']]
    for group in group_fy_data:
        table.append([f"[{group['t_name']}]", group['opening'], group['closing']])
        for acc_type in group['types']:
            table.append([f"  {acc_type['name']}", acc_type['opening'], acc_type['closing']])
            for acc in acc_type['accounts']:
                table.append([f"    {acc['name']}", acc['opening'], acc['closing']])
    table.append(['Total', total_opening, total_closing])
    return table


def build_group_type_summary_table(group_fy_data, total_opening, total_closing):
    """
    Build table_data list for the Group & Type Summary (types level only, no accounts).

    Args:
        group_fy_data (list): Processed group data with nested types.
        total_opening (Decimal): Grand total opening.
        total_closing (Decimal): Grand total closing.

    Returns:
        list[list]: table_data suitable for generate_excel_report / generate_csv_report.
    """
    table = [['Group / Type', 'Opening', 'Closing']]
    for group in group_fy_data:
        table.append([f"[{group['t_name']}]", group['opening'], group['closing']])
        for acc_type in group['types']:
            table.append([f"  {acc_type['name']}", acc_type['opening'], acc_type['closing']])
    table.append(['Total', total_opening, total_closing])
    return table


def build_networth_summary_table(shops, rows):
    """
    Build table_data list for the Networth (Shops Yearly) Summary pivot report.

    Args:
        shops (list): Shop model instances ordered by short_name.
        rows (list): Dicts with keys 'fy_label', 'closings' (list), 'total'.
        col_totals (list): Per-shop column totals.
        grand_total (Decimal): Overall grand total.

    Returns:
        list[list]: table_data suitable for generate_excel_report / generate_csv_report.
    """
    headers = ['FY'] + [s.short_name for s in shops] + ['Total']
    table = [headers]
    for row in rows:
        table.append([row['fy_label']] + list(row['closings']) + [row['total']])
    # table.append(['Total'] + list(col_totals) + [grand_total])
    return table


def generate_excel_report(title, table_data):
    """
    Generate an Excel report with the given title and table data.
    
    Args:
        title (str): The title of the report
        table_data (list): List of lists representing table rows
                          First row should be headers
    
    Returns:
        io.BytesIO: Excel file content as bytes
    """
    excel_buffer = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    
    # Define styles
    header_fill = PatternFill(start_color='1f4788', end_color='1f4788', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=14, color='1f4788')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row counter
    row_num = 1
    
    # Add title
    worksheet.merge_cells(f'A{row_num}:F{row_num}')
    title_cell = worksheet[f'A{row_num}']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    row_num += 1
    
    # Add timestamp
    worksheet.merge_cells(f'A{row_num}:F{row_num}')
    timestamp_cell = worksheet[f'A{row_num}']
    timestamp_cell.value = f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    timestamp_cell.font = Font(italic=True, size=9, color='808080')
    timestamp_cell.alignment = Alignment(horizontal='right', vertical='center')
    row_num += 2
    
    # Add table data
    if table_data:
        for row_data in table_data:
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Convert Decimal to string
                if isinstance(cell_value, Decimal):
                    cell.value = cell_value
                    cell.number_format = '#,##0.00'
                else:
                    cell.value = cell_value
                
                # Apply header styling to first row
                if row_num == 4:  # First data row is row 4 (after title, timestamp)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                
                cell.border = border
            
            row_num += 1
    
    # Adjust column widths
    for col_num in range(1, len(table_data[0]) + 1 if table_data else 1):
        column_letter = chr(64 + col_num)
        worksheet.column_dimensions[column_letter].width = 18
    
    # Save to buffer
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
